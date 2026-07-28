"""
AMI 작업 데이터 → 도로명주소 + 좌표 동시 추출
엑셀(지번/도로명 무관) → ami_data_coords.json

사용 방법:
    python3 좌표추출.py
"""

import openpyxl
import json
import requests
import time
import re
import os
import sys

# 공용 자격증명 로더 (repo 루트 스크립트라 scripts/ 를 경로에 추가)
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scripts'))
from _env import require_env

# ========== 설정 ==========
API_KEY     = require_env('KAKAO_REST_API_KEY')
EXCEL_FILE  = '전기차_마용_도로명.xlsx'
OUTPUT_JSON = 'ami_data_coords.json'
# ==========================


def parse_type(meter_no):
    """계기번호 3~4번째 자리(index 2~3)로 계기타입 판별"""
    code = str(meter_no)[2:4]
    if code == '17':
        return 'E'
    if code == '19':
        return 'EA'
    if code in ('25', '26', '27', '45', '46', '47'):
        return 'G'
    if code in ('53', '55'):
        return 'Amigo'
    return '알수없음'


def fix_meter_no(val):
    """계기번호를 11자리 문자열로 정규화 (float → int → zfill)"""
    if val is None:
        return ''.zfill(11)
    try:
        return str(int(float(val))).zfill(11)
    except (ValueError, TypeError):
        return str(val).strip().zfill(11)


def search_address(query):
    """카카오 주소 검색 API → (도로명주소, lat, lng) 또는 None"""
    url = 'https://dapi.kakao.com/v2/local/search/address.json'
    headers = {'Authorization': f'KakaoAK {API_KEY}'}
    try:
        r = requests.get(url, headers=headers, params={'query': query}, timeout=10)
        if r.status_code == 200:
            docs = r.json().get('documents', [])
            if docs:
                doc = docs[0]
                lat = float(doc['y'])
                lng = float(doc['x'])
                road = (doc.get('road_address') or {}).get('address_name', '')
                jibun = (doc.get('address') or {}).get('address_name', '')
                road_name = road or jibun or query
                return road_name, lat, lng
    except Exception as e:
        print(f"  ⚠️ API 오류: {e}")
    return None


def extract_dong(address):
    """주소에서 '시 구 동/읍/면' 까지만 추출 (폴백용)"""
    m = re.match(r'(.*?[동읍면])', address)
    return m.group(1).strip() if m else None


def main():
    print("=" * 60)
    print("📍 AMI 데이터 도로명주소 + 좌표 추출")
    print("=" * 60)

    # 엑셀 읽기
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"컬럼: {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        rows.append(d)
    print(f"총 {len(rows)}개 행 읽음\n")

    # 주소별로 그룹핑 (API 호출 최소화)
    grouped = {}
    for d in rows:
        addr = str(d.get('고객주소') or '').strip()
        if addr not in grouped:
            grouped[addr] = []
        grouped[addr].append(d)

    result = []
    total = len(grouped)
    exact_cnt = 0
    approx_cnt = 0
    fail_cnt = 0

    for i, (addr, items) in enumerate(grouped.items(), 1):
        road_addr = str(items[0].get('도로명주소') or '').strip()

        # 1차: 도로명주소로 시도
        found = None
        if road_addr:
            found = search_address(road_addr)
            if found:
                accuracy = 'exact'

        # 2차: 지번주소로 재시도
        if not found and addr:
            found = search_address(addr)
            if found:
                accuracy = 'exact'

        # 3차: 동 이름으로 폴백
        if not found:
            dong = extract_dong(addr) or extract_dong(road_addr)
            if dong:
                found = search_address(dong)
                if found:
                    accuracy = 'approximate'

        if found:
            resolved_road, lat, lng = found
            flag = '✅' if accuracy == 'exact' else '🟡'
            print(f"[{i}/{total}] {flag} {addr}")
            if accuracy == 'approximate':
                dong_used = extract_dong(addr) or extract_dong(road_addr)
                print(f"  └ 폴백({dong_used}) → {lat:.6f}, {lng:.6f}")

            for d in items:
                meter_no = fix_meter_no(d.get('계기번호'))
                상호_val = d.get('상호')
                상호 = str(상호_val) if 상호_val and 상호_val != 0 and str(상호_val) != '0' else ''
                result.append({
                    '주소': addr,
                    '도로명주소': resolved_road,
                    '계기번호': meter_no,
                    '계기타입': parse_type(meter_no),
                    '변대주': str(d.get('변대주전산화번호') or ''),
                    '상호': 상호,
                    'lat': lat,
                    'lng': lng,
                    '좌표정확도': accuracy,
                })

            if accuracy == 'exact':
                exact_cnt += 1
            else:
                approx_cnt += 1
        else:
            print(f"[{i}/{total}] ❌ {addr} (완전 실패)")
            fail_cnt += 1

        time.sleep(0.15)

    # JSON 저장
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print()
    print("=" * 60)
    print("✅ 완료!")
    print(f"  정확  (exact)      : {exact_cnt}개 주소")
    print(f"  근사  (approximate): {approx_cnt}개 주소")
    print(f"  실패               : {fail_cnt}개 주소")
    print(f"  계기 총계          : {len(result)}개")
    print(f"📁 저장: {OUTPUT_JSON}")
    print("=" * 60)


if __name__ == '__main__':
    main()
