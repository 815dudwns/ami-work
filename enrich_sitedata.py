"""
메인 엑셀(계기교체_보강현황)에서 site-data.json을 보강한다.

원칙:
- 계기번호로 매칭
- 엑셀에 있는 컬럼만 동적으로 가져온다 (한전이 매번 다르게 줄 수 있음)
- 지번/도로명/공동주택명/상호명은 이미 다른 곳에 표시되므로 제외
- 18~22번 빈 헤더 컬럼도 제외
- 엑셀 값이 비어있거나 #N/A 면 기존값 유지
- 신/구 같이 들어오면 신(컬럼 23~26)이 우선
"""
import json
import shutil
from datetime import datetime
from openpyxl import load_workbook

EXCEL = '/Users/woodelight/Projects/ami-work/계기교체_보강현황_20260513.xlsx'
SITE = '/Users/woodelight/Projects/ami-work/data/site-data.json'
SHEET = '계기교체 보강현황_20260513155458'

# 제외할 컬럼 (이미 다른 곳에 표시되거나 빈 헤더)
EXCLUDE = {'지번', '도로명', '공동주택명', '상호명'}

# 엑셀 컬럼명 → JSON 키 매핑 (없으면 엑셀 이름 그대로 사용)
# 신/구 같은 이름이면 인덱스로 구분
RENAME = {
    'DCU ID': 'DCUID',
    '모뎀 MAC': '모뎀MAC',
    '시스템등록일(영배)': '시스템등록일',
    '계기교체일(A)': '계기교체일',
    '연계 수신일(B)': '연계수신일',
    '최초LP 수신일(C)': '최초LP수신일',
    '시스템등록 소요일(B)-(A)': '등록소요일',
    'DCU 장애여부': 'DCU장애여부',
}


def norm(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if s in ('', '#N/A', 'nan', 'None'):
        return ''
    # 숫자 float → int 처리 (계기번호 등)
    if s.endswith('.0') and s[:-2].isdigit():
        return s[:-2]
    return s


def main():
    # site-data 백업
    ts = datetime.now().strftime('%Y%m%d-%H%M%S')
    backup = SITE + f'.backup-{ts}'
    shutil.copy2(SITE, backup)
    print(f"[backup] {backup}")

    print(f"[load] {EXCEL}")
    wb = load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb[SHEET]

    # 헤더 분석 — 같은 이름이 두 번 나오면 _전 / (신은 기본)
    raw_header = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        raw_header = list(row)
        break

    # 첫 등장은 "_전", 두 번째 등장은 기본 이름으로
    seen = {}
    headers = []
    for idx, h in enumerate(raw_header):
        if h is None or str(h).strip() == '':
            headers.append(None)
            continue
        h = str(h).strip()
        if h in seen:
            # 두 번째 등장 — 기본 이름 그대로 (신 정보)
            headers.append(h)
        else:
            # 첫 등장 — 다음에 또 나오면 "_전" 처리해야 함
            seen[h] = idx
            headers.append(h)

    # 두 번 등장한 이름을 찾아 첫 등장만 "_전" 으로 바꾸기
    name_count = {}
    for h in headers:
        if h:
            name_count[h] = name_count.get(h, 0) + 1
    duplicates = {h for h, c in name_count.items() if c > 1}

    final_headers = []
    seen2 = set()
    for h in headers:
        if h is None:
            final_headers.append(None)
        elif h in duplicates and h not in seen2:
            # 첫 등장 → _전
            final_headers.append(h + '_전')
            seen2.add(h)
        else:
            final_headers.append(h)

    # 제외/리네임 적용
    json_keys = []
    for h in final_headers:
        if h is None or h in EXCLUDE:
            json_keys.append(None)
        elif h in RENAME:
            json_keys.append(RENAME[h])
        else:
            json_keys.append(h)

    # 계기번호 컬럼 위치
    try:
        meter_col = final_headers.index('계기번호')
    except ValueError:
        raise RuntimeError("'계기번호' 컬럼을 못 찾음")

    # 엑셀 → dict (계기번호 → {key: value})
    excel_map = {}
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        mtr = norm(row[meter_col])
        if not mtr:
            continue
        record = {}
        for i, val in enumerate(row):
            key = json_keys[i]
            if key is None:
                continue
            v = norm(val)
            if v:
                record[key] = v
        excel_map[mtr] = record

    wb.close()
    print(f"[excel] {row_count:,} 행, {len(excel_map):,} 고유 계기번호")
    print(f"[fields] 가져올 키: {sorted({k for r in excel_map.values() for k in r})}")

    # site-data.json 로드 + 머지
    sd = json.load(open(SITE, encoding='utf-8'))
    print(f"[json] 현재 {len(sd):,} 항목")

    matched = 0
    enriched_fields = {}  # key → 추가/덮어쓴 카운트
    for item in sd:
        mtr = norm(item.get('계기번호'))
        if not mtr or mtr not in excel_map:
            continue
        matched += 1
        rec = excel_map[mtr]
        for k, v in rec.items():
            existing = norm(item.get(k))
            if v != existing:
                item[k] = v
                enriched_fields[k] = enriched_fields.get(k, 0) + 1

    print(f"[merge] 매칭된 항목: {matched:,} / {len(sd):,}")
    print(f"[merge] 필드별 변경 카운트:")
    for k in sorted(enriched_fields, key=lambda x: -enriched_fields[x]):
        print(f"   {k:<14} {enriched_fields[k]:>6,}")

    # 저장
    with open(SITE, 'w', encoding='utf-8') as f:
        json.dump(sd, f, ensure_ascii=False, indent=2)
    print(f"[saved] {SITE}")


if __name__ == '__main__':
    main()
