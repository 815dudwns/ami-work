#!/usr/bin/env python3
"""고압철거 데이터셋에 주덕기 과장 '주소추가' 시트분을 반영한다.

원천 = data/inbox_jdg_20260827/25년보강_고압철거대상_v3.xlsx 의 **주소추가 시트**.
  (주덕기 KDN 과장 2026-08-27 08:32 메일 "주소 추가했습니다." 첨부)
  메일 지시: "주소추가 시트 확인 하시기 바랍니다. 4개 미입력부분있는거 제외하고 반영하시면 됩니다."

주소추가 시트 = Sheet2 429행 중 이번 v3 에서 주소가 새로 채워진 76행의 발췌본이다.
  76행 전부 Sheet2 에도 있고, 기존 gapap-data.json 168건과는 계기번호2 기준 중복 0건이다.

작업대상 필터 — apply_gapap_sheet2_fields.py 와 같은 3조건(2026-08-18 영준님 확정):
  1. 한전기준 이 있을 것
  2. 주소 가 있을 것        -> 4건 제외 (메일이 말한 "4개 미입력")
  3. 대상 == '철거'         -> 1건 제외 (대상=정상)
  결과 71건 = 한전기준 '철거+재설치' 69 + '철거' 2.

계기번호는 **계기번호2** 를 쓴다(2026-08-18 영준님 확정). 기존 168건이 전부 계기번호2 로
  저장돼 있는 것을 확인했다(168/168). 계기번호1 은 데이터에 남기지 않고, 고객번호·MAC 을
  Sheet1 에서 끌어올 때의 조회키로만 쓴다(Sheet1 은 계기번호1 기준, 71/71 매칭).

고압철거는 전주 변대주 계통이 아니라 건물 고압변압기 모자분리 계기라 **DCU 매칭을 하지 않는다**.
  DCUID/변대주/dcu_철거예정 은 빈 문자열로 두고 apply_dcu_status.py 를 돌리지 않는다.

주소 표기가 v2 와 다르다: v3 주소추가는 `도로명  (법정동 번지 상세)` 꼴이다(71건 중 60건).
  `주소` 에는 원문을 그대로 싣고, 괄호 안의 법정동+번지를 지오코딩 지번 질의로 따로 만든다.

사용:
    python3 scripts/apply_gapap_juso_chuga.py            # 예행(dry-run) — 파일 안 씀
    python3 scripts/apply_gapap_juso_chuga.py --apply     # 실제 반영(+백업)
"""

import argparse
import collections
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from geocode_cascade import resolve  # noqa: E402

TARGET = ROOT / "data" / "gapap-data.json"
XLSX = ROOT / "data" / "inbox_jdg_20260827" / "25년보강_고압철거대상_v3.xlsx"
SHEET = "주소추가"
KST = ZoneInfo("Asia/Seoul")

EXPECT_TOTAL = 71          # 필터 통과 예상 건수
EXPECT_BEFORE = 168        # 반영 전 gapap-data.json 건수

# 계기타입 재판정 — 계기번호 3~4번째 자리 코드만 본다(CLAUDE.md 데이터규칙).
#   apply_gapap_sheet2_fields.py 와 같은 표를 쓴다. 어휘도 이 데이터셋 표기(G타입/AE타입/...)를 따른다.
TYPE_BY_CODE = {
    "17": "E타입",
    "19": "AE타입",
    "25": "G타입", "26": "G타입", "27": "G타입",
    "45": "G타입", "46": "G타입", "47": "G타입",
    "53": "보안계기", "55": "보안계기",
}
# 코드가 매핑에 없으면(38·52 등) 지어내지 않고 엑셀 계기타입을 그대로 쓴다.
TYPE_BY_XLS = {
    "E": "E타입",
    "AE": "AE타입", "EA": "AE타입",
    "G": "G타입",
    "AMIGO": "보안계기", "보안계기": "보안계기",
}

EMPTY_TOKENS = {"", "0", "#N/A", "#n/a", "nan", "NaN", "None", "none", "-"}

# gapap-data.json 레코드의 필드 순서 — 기존 168건과 같은 모양으로 만든다.
FIELD_ORDER = [
    "지사", "주소", "도로명주소", "계기번호", "계기타입", "고객번호", "통신방식",
    "공동주택명", "상호", "검기만료년월", "계기타입_전", "인입주", "변대주", "DCUID",
    "lat", "lng", "좌표정확도", "DCU장애여부", "교체사유", "시스템등록일", "계기교체일",
    "연계수신일", "등록소요일", "사업차수_전", "통신방식_전", "검침방법_전", "검침방법",
    "사업차수", "모뎀MAC", "비고", "dcu_철거예정", "한전기준", "비고1", "비고2", "주소2",
]


def norm(v):
    """셀 값 정규화 — 빈 의미의 토큰은 전부 빈 문자열로."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return "" if s in EMPTY_TOKENS else s


def meter_no(v):
    """계기번호 정규화 — float -> int -> str -> 하이픈 제거 -> zfill(11)."""
    if v is None:
        return ""
    if isinstance(v, float):
        v = int(v)
    s = str(v).strip().replace("-", "")
    if not s or s.lower() == "nan":
        return ""
    return s.zfill(11)


def meter_type(no, xls_type):
    """계기타입 재판정. 코드 미매핑이면 엑셀 값을 따른다(지어내지 않는다)."""
    code = no[2:4] if len(no) >= 4 else ""
    if code in TYPE_BY_CODE:
        return TYPE_BY_CODE[code]
    key = (xls_type or "").strip().upper()
    return TYPE_BY_XLS.get(key, xls_type or "")


def read_sheet(wb, name):
    """헤더 중복(계기번호가 두 번 나오는 Sheet1)을 견디며 행을 dict 로 읽는다."""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    out = []
    for r in rows[1:]:
        if not any(x is not None and str(x).strip() != "" for x in r):
            continue
        d = {}
        for i, k in enumerate(header):
            if k is None:
                continue
            key = k if k not in d else f"{k}_2"
            d[key] = r[i]
        out.append(d)
    return out


# 원천 주소 오타 교정 — 근거를 적고 건별로만 연다. 일괄 추정 교정은 하지 않는다.
#   ★`주소` 필드까지 교정한다(영준님 지시 2026-08-27 "연세로로 고쳐"). 지오코딩 질의만
#     바꾸고 원문을 남겨 두면 지도 주소에 없는 도로가 그대로 뜬다.
#     교정 전 원문은 이 표와 커밋 메시지에 남는다.
ADDR_FIX = {
    # 계기번호2: (도로명질의, 지번질의, 사유)
    "79450102396": (
        "서울특별시 서대문구 연세로5다길 10",
        "서울특별시 서대문구 창천동 62-11",
        "원천 '서대문구 연서로5다길 10' — 연서로는 은평구 도로이고 서대문구에 없다. "
        "같은 행 비고2 '기존계기교체 연세로5가길8-3' 과 같은 리스트의 79450102426 "
        "'서대문구 연세로5다길 10 (창천동 62-11)' 이 같은 건물을 가리킨다. 연세로 오타로 본다.",
    ),
}


def _looks_like_road(s):
    """'...로 12' / '...길 8' 처럼 도로명+건물번호 꼴인지. 아니면 지번 표기로 본다."""
    return bool(re.search(r"(로|길)\s*\d", s))


def split_addr(raw):
    """'도로명  (법정동 번지 상세)' -> (도로명질의, 지번질의).

    괄호가 없거나 괄호 안에 법정동이 없으면 지번질의는 빈 문자열이고 도로명만 쓴다.
    지오코딩은 도로명 -> 지번 순으로 캐스케이드하므로 지번이 없어도 동작한다.
    ★괄호 앞이 도로명이 아니라 지번인 행이 있다(예: '용산구 이촌동 196-3 (철2)한강철교...').
      그대로 road 로 넘기면 네이버 근접응답을 타 approximate 가 된다. 지번으로 넘긴다.
    """
    s = re.sub(r"\s+", " ", raw or "").strip()
    m = re.match(r"^(.*?)\s*\((.*)\)\s*$", s)
    if not m:
        return (s, "") if _looks_like_road(s) else ("", s)
    road = m.group(1).strip()
    inner = m.group(2).strip()
    if not _looks_like_road(road):
        return "", road

    # 지번질의에 붙일 '시 구' 접두. 도로명 앞머리에서 뽑는다.
    pm = re.match(r"^(\S+?(?:특별시|광역시|특별자치시|도))\s+(\S+?(?:구|시|군))\b", road)
    if not pm:
        return road, ""
    prefix = f"{pm.group(1)} {pm.group(2)}"

    # 괄호 안 첫 토큰이 법정동이어야 한다. '( 172번지)' 처럼 동이 없으면 포기한다.
    dm = re.match(r"^(\S*?[동가읍면리])\s*(\d+(?:-\d+)?)(?:번지)?", inner)
    if not dm:
        return road, ""
    dong, beonji = dm.group(1), dm.group(2)
    beonji = re.sub(r"-0$", "", beonji)   # '274-0' 은 부번 없음을 뜻한다
    return road, f"{prefix} {dong} {beonji}"


def build():
    wb = load_workbook(XLSX, data_only=True)
    add_rows = read_sheet(wb, SHEET)
    s1_rows = read_sheet(wb, "Sheet1")

    # Sheet1 조회표 — 계기번호1 기준(고객번호·MAC 출처)
    s1 = {}
    for r in s1_rows:
        key = meter_no(r.get("계기번호"))
        if key and key not in s1:
            s1[key] = r

    skipped = collections.defaultdict(list)
    work = []
    for r in add_rows:
        addr = norm(r.get("주소"))
        target = norm(r.get("대상"))
        rule = norm(r.get("한전기준"))
        if not addr:
            skipped["주소 미입력"].append(r)
            continue
        if target != "철거":
            skipped[f"대상={target or '(빈칸)'}"].append(r)
            continue
        if not rule:
            skipped["한전기준 없음"].append(r)
            continue
        work.append(r)

    return work, skipped, s1


def make_record(r, s1, geo_cache):
    m1 = meter_no(r.get("계기번호"))       # 조회용(Sheet1 키). 데이터에는 남기지 않는다.
    m2 = meter_no(r.get("계기번호2")) or m1
    src = s1.get(m1, {})

    raw_addr = re.sub(r"\s+", " ", norm(r.get("주소"))).strip()
    if m2 in ADDR_FIX:
        road_q, jibun_q, _why = ADDR_FIX[m2]
        raw_addr = road_q          # 주소 원문도 교정본으로 바꾼다
    else:
        road_q, jibun_q = split_addr(raw_addr)

    ck = (road_q, jibun_q)
    if ck not in geo_cache:
        geo_cache[ck] = resolve(jibun=jibun_q, road=road_q)
    hit = geo_cache[ck]

    bigo1, bigo2 = norm(r.get("비고1")), norm(r.get("비고2"))
    bigo = " / ".join(x for x in (bigo1, bigo2) if x)

    mac = norm(r.get("MAC")) or norm(src.get("MAC"))

    rec = {
        "지사": norm(r.get("2차사업소")),
        "주소": raw_addr,
        "도로명주소": (hit.road or road_q) if hit.accuracy != "fail" else road_q,
        "계기번호": m2,
        "계기타입": meter_type(m2, norm(r.get("계기타입"))),
        "고객번호": norm(src.get("고객번호")),
        "통신방식": norm(r.get("통신방식")) or "LTE",
        "공동주택명": "",
        "상호": "",
        "검기만료년월": "",
        "계기타입_전": "",
        "인입주": "",
        "변대주": "",          # 고압철거는 전주 계통이 아니다 — 비운다
        "DCUID": "",           # 같은 이유로 비운다(DCU 매칭 금지)
        "lat": hit.lat if hit.accuracy != "fail" else None,
        "lng": hit.lng if hit.accuracy != "fail" else None,
        "좌표정확도": hit.accuracy,
        "DCU장애여부": "",
        "교체사유": "고압철거",
        "시스템등록일": "",
        "계기교체일": "",
        "연계수신일": "",
        "등록소요일": "",
        "사업차수_전": "",
        "통신방식_전": "",
        "검침방법_전": "",
        "검침방법": "",
        "사업차수": "",
        "모뎀MAC": mac,
        "비고": bigo,
        "dcu_철거예정": "",
        "한전기준": norm(r.get("한전기준")),
        "비고1": bigo1,
        "비고2": bigo2,
        "주소2": norm(r.get("주소2")),
    }
    # 기존 168건과 같은 키 순서로 정렬
    return {k: rec[k] for k in FIELD_ORDER}, (m1, m2, mac, norm(src.get("MAC")))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="실제 파일에 반영(+백업)")
    args = ap.parse_args()

    work, skipped, s1 = build()

    print(f"[원천] {XLSX.name} / 시트 '{SHEET}'")
    print(f"[필터] 작업대상 {len(work)}건")
    for reason, rows in skipped.items():
        print(f"   제외 {reason}: {len(rows)}건")
        for r in rows:
            print(f"      - {norm(r.get('2차사업소'))} | {norm(r.get('주소')) or '(주소없음)'} "
                  f"| 계기 {meter_no(r.get('계기번호'))} | 비고1 {norm(r.get('비고1'))}")

    if len(work) != EXPECT_TOTAL:
        print(f"[중단] 작업대상이 예상({EXPECT_TOTAL})과 다르다: {len(work)}")
        return 1

    cur = json.loads(TARGET.read_text(encoding="utf-8"))
    print(f"[대상] {TARGET.name} 현재 {len(cur)}건")
    if len(cur) != EXPECT_BEFORE:
        print(f"[중단] 반영 전 건수가 예상({EXPECT_BEFORE})과 다르다: {len(cur)}")
        return 1

    geo_cache = {}
    records, audit = [], []
    for i, r in enumerate(work, 1):
        rec, info = make_record(r, s1, geo_cache)
        records.append(rec)
        audit.append(info)
        print(f"   {i:3d}/{len(work)} {rec['지사']:8s} {rec['계기번호']} "
              f"{rec['계기타입']:6s} {rec['한전기준']:7s} {rec['좌표정확도']:12s} {rec['주소'][:44]}")

    # --- 검증 ---
    cur_nos = {meter_no(x.get("계기번호")) for x in cur}
    dup = [x["계기번호"] for x in records if x["계기번호"] in cur_nos]
    if dup:
        print(f"[중단] 기존과 계기번호 중복 {len(dup)}건: {dup[:10]}")
        return 1

    inner_dup = [n for n, c in collections.Counter(x["계기번호"] for x in records).items() if c > 1]
    if inner_dup:
        print(f"[중단] 신규분 내부 계기번호 중복: {inner_dup}")
        return 1

    no_cust = [x["계기번호"] for x in records if not x["고객번호"]]
    no_mac = [x["계기번호"] for x in records if not x["모뎀MAC"]]
    mac_mismatch = [(m2, a, b) for (m1, m2, a, b) in audit if a and b and a != b]
    fails = [x for x in records if x["좌표정확도"] == "fail"]
    approx = [x for x in records if x["좌표정확도"] == "approximate"]

    print("\n===== 요약 =====")
    print(f"신규 {len(records)}건 -> 합계 {len(cur) + len(records)}건")
    print("지사별:", dict(collections.Counter(x["지사"] for x in records)))
    print("한전기준:", dict(collections.Counter(x["한전기준"] for x in records)),
          "(철거+재설치=마커 '교' / 철거=마커 '철')")
    print("계기타입:", dict(collections.Counter(x["계기타입"] for x in records)))
    print("좌표:", dict(collections.Counter(x["좌표정확도"] for x in records)))
    for x in approx:
        print(f"   approximate> {x['지사']} | {x['주소']}")
    for x in fails:
        print(f"   FAIL> {x['지사']} | {x['주소']}")
    # 고객번호는 Sheet1 6184행 중 3370행에만 있고, 이번 71건에는 한 건도 없다.
    #   기존 168건도 67건만 갖고 있어 빈칸이 이 데이터셋의 정상 상태다(막지 않는다).
    print(f"고객번호 미매칭: {len(no_cust)}건 (기존 168건 중에도 101건이 빈칸 — 정상)")
    if ADDR_FIX:
        print("주소 오타 교정(지오코딩 질의만, `주소` 원문은 보존):")
        for k, (rd, jb, why) in ADDR_FIX.items():
            if any(x["계기번호"] == k for x in records):
                print(f"   {k} -> 도로명 '{rd}' / 지번 '{jb}'\n      사유: {why}")
    print(f"MAC 없음: {len(no_mac)}건 {no_mac[:10]}")
    print(f"MAC 불일치(주소추가 vs Sheet1): {len(mac_mismatch)}건")
    for m2, a, b in mac_mismatch[:10]:
        print(f"   {m2}: 주소추가={a} / Sheet1={b}  -> 주소추가 값 채택")

    if not args.apply:
        print("\n[예행] --apply 를 붙이면 실제로 씁니다.")
        return 0

    ts = datetime.now(KST).strftime("%Y%m%d-%H%M%S")
    backup = TARGET.with_name(f"gapap-data.backup-주소추가{len(records)}건전-{ts}.json")
    shutil.copy2(TARGET, backup)
    print(f"\n[백업] {backup.name}")

    merged = cur + records
    # 기존 파일이 압축 JSON(들여쓰기 없음, 기본 구분자)이라 그 형식을 그대로 따른다.
    TARGET.write_text(json.dumps(merged, ensure_ascii=False), encoding="utf-8")
    print(f"[반영] {TARGET.name} {len(cur)} -> {len(merged)}건")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
