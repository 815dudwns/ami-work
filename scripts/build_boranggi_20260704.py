#!/usr/bin/env python3
"""보강현황 신규 계기 site-data 레코드 생성 + 좌표변환 (트랙1)
입력:
  - data/boranggi-20260703.xlsx  (새 엑셀)
  - data/site-data.backup-완료제거전-20260704.json  (원본 26,588건 — 신규 판정 기준)
출력:
  - data/boranggi-new-20260704.json  (신규 레코드 배열, site-data.json 직접수정 금지)

처리:
  1. xlsx unzip → sheet1.xml → hidden="1" 숨김행 추출 → 표시행(활성행)만 파싱
  2. 백업 site-data 고객번호 집합으로 신규 판정 (zfill10 정규화 필수)
  3. 백업 좌표 시드: (지번, 도로명) 동일하면 기존 lat/lng/정확도 재사용 → 나머지만 카카오→네이버→동중심 지오코딩
  4. 계기타입: 계기번호 [2:4] 파싱 (기존 vocab 준수)
  5. 무결성 자체검증
"""
import json
import re
import sys
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

BASE = Path('/Users/woodelight/Projects/ami-work')
EXCEL = BASE / 'data/boranggi-20260703.xlsx'
BACKUP_SITE = BASE / 'data/site-data.backup-완료제거전-20260704.json'
OUT_JSON = BASE / 'data/boranggi-new-20260704.json'

from _env import require_env

KAKAO_KEY = require_env('KAKAO_REST_API_KEY')
NAVER_ID = require_env('NAVER_GEOCODE_ID')
NAVER_SECRET = require_env('NAVER_GEOCODE_SECRET')

WORKERS = 10

# 계기타입 매핑 (기존 site-data vocab 준수)
METER_TYPE_MAP = {
    '17': 'E',          # CLAUDE.md 기준 신규 (3건)
    '19': 'AE타입',     # 기존 vocab
    '25': 'G타입', '26': 'G타입', '27': 'G타입',
    '45': 'G타입', '46': 'G타입', '47': 'G타입',
    '53': '보안계기', '55': '보안계기',
}

# ─── 정규화 함수 ──────────────────────────────────────────────────────────────

def normalize_meter(v):
    """계기번호: float→int→str→zfill(11), 알파포함이면 as-is"""
    s = str(v).strip() if v is not None else ''
    if not s:
        return ''
    if s.replace('.', '', 1).lstrip('-').isdigit():
        return str(int(float(s))).zfill(11)
    return s


def normalize_cust_no(v):
    """고객번호: float→int→str→zfill(10)"""
    s = str(v).strip() if v is not None else ''
    if not s:
        return ''
    if s.replace('.', '', 1).lstrip('-').isdigit():
        return str(int(float(s))).zfill(10)
    return s


def normalize_str(v):
    """문자열: None/'0'/'#N/A'/'None' → ''"""
    if v is None:
        return ''
    s = str(v).strip()
    if s in ('0', '#N/A', 'None', '#VALUE!', '#REF!'):
        return ''
    return s


def normalize_date(v):
    """날짜: datetime → 'YYYY-MM-DD', 그 외 → ''"""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if s in ('', 'None', '#N/A'):
        return ''
    return s


def parse_meter_type(meter_no):
    """계기번호 [2:4]에서 계기타입 파싱"""
    if len(meter_no) < 4:
        return '알수없음'
    prefix = meter_no[2:4]
    return METER_TYPE_MAP.get(prefix, '알수없음')


# ─── 숨김행 추출 ─────────────────────────────────────────────────────────────

def get_hidden_row_set(xlsx_path):
    """sheet1.xml hidden="1" 패턴으로 숨김행 번호 집합 반환 (1-base)"""
    with zipfile.ZipFile(xlsx_path) as z:
        with z.open('xl/worksheets/sheet1.xml') as f:
            content = f.read()
    hidden = set(int(m) for m in re.findall(rb'<row r="(\d+)"[^>]*hidden="1"', content))
    return hidden


# ─── 좌표 추출 ───────────────────────────────────────────────────────────────

_kakao = requests.Session()
_kakao.headers.update({'Authorization': f'KakaoAK {KAKAO_KEY}'})

_naver = requests.Session()
_naver.headers.update({
    'x-ncp-apigw-api-key-id': NAVER_ID,
    'x-ncp-apigw-api-key': NAVER_SECRET,
})


def kakao_search_address(query):
    if not query:
        return None
    for attempt in range(4):
        try:
            r = _kakao.get(
                'https://dapi.kakao.com/v2/local/search/address.json',
                params={'query': query}, timeout=15)
            if r.status_code == 429:
                time.sleep(0.5 * (attempt + 1))
                continue
            if r.status_code == 200:
                docs = r.json().get('documents', [])
                if docs:
                    doc = docs[0]
                    road = (doc.get('road_address') or {}).get('address_name', '')
                    jibun = (doc.get('address') or {}).get('address_name', '')
                    return float(doc['y']), float(doc['x']), (road or jibun or query)
            time.sleep(0.3)
        except Exception:
            time.sleep(0.3)
    return None


def kakao_search_keyword(query):
    if not query:
        return None
    try:
        r = _kakao.get(
            'https://dapi.kakao.com/v2/local/search/keyword.json',
            params={'query': query, 'size': 1}, timeout=15)
        if r.status_code == 200:
            docs = r.json().get('documents', [])
            if docs:
                doc = docs[0]
                addr = doc.get('road_address_name') or doc.get('address_name') or query
                return float(doc['y']), float(doc['x']), addr
    except Exception:
        pass
    return None


def naver_geocode(query):
    """네이버 NCP Geocoding"""
    if not query:
        return None
    try:
        r = _naver.get(
            'https://maps.apigw.ntruss.com/map-geocode/v2/geocode',
            params={'query': query}, timeout=15)
        if r.status_code == 200:
            data = r.json()
            addresses = data.get('addresses', [])
            if addresses:
                a = addresses[0]
                return float(a['y']), float(a['x']), a.get('roadAddress') or query
    except Exception:
        pass
    return None


def extract_dong(address):
    m = re.match(r'(.*?[동읍면])', address or '')
    return m.group(1).strip() if m else None


def trim_addr(addr):
    if not addr:
        return ''
    a = re.sub(r'\s+', ' ', addr).strip()
    return re.split(r'[()]', a)[0].strip()


def resolve_coords(jibun_raw, road_raw):
    """주소 → (accuracy, lat, lng)
    cascade: 카카오주소(도로명) → 카카오주소(지번) → 카카오키워드(도로명) → 카카오키워드(지번)
             → 네이버(도로명) → 네이버(지번) → 동중심
    """
    jibun = trim_addr(jibun_raw)
    road = trim_addr(road_raw)

    # 1. 카카오 도로명
    if road:
        r = kakao_search_address(road)
        if r:
            return 'exact', r[0], r[1]

    # 2. 카카오 지번
    if jibun:
        r = kakao_search_address(jibun)
        if r:
            return 'exact', r[0], r[1]

    # 3. 카카오 키워드 도로명
    if road:
        r = kakao_search_keyword(road)
        if r:
            return 'exact', r[0], r[1]

    # 4. 카카오 키워드 지번
    if jibun:
        r = kakao_search_keyword(jibun)
        if r:
            return 'exact', r[0], r[1]

    # 5. 네이버 도로명
    if road:
        r = naver_geocode(road)
        if r:
            return 'exact', r[0], r[1]

    # 6. 네이버 지번
    if jibun:
        r = naver_geocode(jibun)
        if r:
            return 'exact', r[0], r[1]

    # 7. 동 중심 폴백
    dong = extract_dong(jibun) or extract_dong(road)
    if dong:
        r = kakao_search_address(dong) or kakao_search_keyword(dong) or naver_geocode(dong)
        if r:
            return 'approximate', r[0], r[1]

    return 'fail', None, None


# ─── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    # 1. 백업 site-data 로드 (신규 판정 기준)
    print('백업 site-data 로드 중...', flush=True)
    with open(BACKUP_SITE) as f:
        backup = json.load(f)
    existing_cust = set()
    for e in backup:
        c = normalize_cust_no(e.get('고객번호', ''))
        if c:
            existing_cust.add(c)
    print(f'백업 고객번호 집합: {len(existing_cust):,}건', flush=True)

    # 좌표 시드: (지번, 도로명) → (lat, lng, 정확도) — 백업에서 추출
    coord_seed = {}
    for e in backup:
        lat = e.get('lat')
        lng = e.get('lng')
        acc = e.get('좌표정확도')
        if lat and lng:
            key = (normalize_str(e.get('주소', '')), normalize_str(e.get('도로명주소', '')))
            if key not in coord_seed:
                coord_seed[key] = (lat, lng, acc)
    print(f'좌표 시드 캐시: {len(coord_seed):,}건', flush=True)

    # 2. 숨김행 번호 집합
    print('숨김행 분석 중...', flush=True)
    hidden_rows = get_hidden_row_set(EXCEL)
    print(f'숨김행: {len(hidden_rows):,}건', flush=True)

    # 3. 엑셀 파싱 (활성행만)
    print('엑셀 파싱 중...', flush=True)
    wb = load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb.active

    new_records = []
    meter_type_dist = {}
    skipped_existing = 0
    skipped_hidden = 0
    skipped_no_meter = 0

    for row in ws.iter_rows(min_row=2):
        row_num = row[0].row
        if row_num in hidden_rows:
            skipped_hidden += 1
            continue

        vals = [c.value for c in row]

        # 계기번호 정규화
        meter_no = normalize_meter(vals[20]) if len(vals) > 20 else ''
        if not meter_no:
            skipped_no_meter += 1
            continue

        # 고객번호 정규화
        cust_no = normalize_cust_no(vals[3]) if len(vals) > 3 else ''

        # 신규 판정: 고객번호가 백업에 없는 것
        if cust_no and cust_no in existing_cust:
            skipped_existing += 1
            continue

        # 계기타입 파싱
        meter_type = parse_meter_type(meter_no)
        meter_type_dist[meter_type] = meter_type_dist.get(meter_type, 0) + 1

        # 컬럼값 추출 (새 엑셀 인덱스 기준)
        # 0지사 3고객번호 4DCU장애여부 5교체사유 6시스템등록일 7검기만료년월 8계기교체일
        # 9연계수신일 10최초LP수신일 15DCUID 16통신방식 17검침방법 19모뎀MAC
        # 20계기번호 21계기타입(신뢰X) 30변대주 31인입주 32지번 33도로명 34공동주택명 35상호

        def g(idx):
            return vals[idx] if len(vals) > idx else None

        jibun = normalize_str(g(32))
        road = normalize_str(g(33))
        biandaeju = normalize_str(g(30))
        ibipju = normalize_str(g(31))
        sangho = normalize_str(g(35))

        # 좌표 시드 재사용 여부 (주소 동일하면 기존 좌표 재사용)
        coord_key = (jibun, road)
        if coord_key in coord_seed:
            lat, lng, acc = coord_seed[coord_key]
        else:
            lat, lng, acc = None, None, None  # 나중에 지오코딩

        record = {
            '지사': normalize_str(g(0)),
            '주소': jibun,
            '도로명주소': road,
            '계기번호': meter_no,
            '계기타입': meter_type,
            '고객번호': cust_no,
            '통신방식': normalize_str(g(16)),
            '공동주택명': normalize_str(g(34)),
            '상호': sangho,
            '검기만료년월': normalize_str(g(7)),
            '계기타입_전': '',
            '인입주': ibipju,
            '변대주': biandaeju,
            'DCUID': normalize_str(g(15)),
            'lat': lat,
            'lng': lng,
            '좌표정확도': acc,
            'DCU장애여부': normalize_str(g(4)),
            '교체사유': normalize_str(g(5)),
            '시스템등록일': normalize_date(g(6)),
            '계기교체일': normalize_date(g(8)),
            '연계수신일': normalize_date(g(9)),
            '등록소요일': str(g(12)) if g(12) is not None else '',
            '사업차수_전': '',  # 새 엑셀 col27은 사업차수이나 블록B(무시). 재사용X
            '통신방식_전': normalize_str(g(16)),  # 현재값 보존
            '검침방법_전': normalize_str(g(17)),
            '검침방법': normalize_str(g(17)),
            '사업차수': '',
        }
        new_records.append(record)

    wb.close()
    print(f'\n파싱 완료:', flush=True)
    print(f'  활성행 처리: {len(new_records) + skipped_existing:,}건', flush=True)
    print(f'  기존 제외: {skipped_existing:,}건', flush=True)
    print(f'  신규: {len(new_records):,}건', flush=True)
    print(f'  계기번호 없음 건너뜀: {skipped_no_meter}건', flush=True)
    print(f'계기타입 분포: {meter_type_dist}', flush=True)

    # 4. 좌표 미지정 신규 레코드 지오코딩
    needs_geocoding = [r for r in new_records if r['lat'] is None]
    seeded = len(new_records) - len(needs_geocoding)
    print(f'\n좌표 시드 재사용: {seeded:,}건', flush=True)
    print(f'지오코딩 필요: {len(needs_geocoding):,}건', flush=True)

    if needs_geocoding:
        # 고유 주소 그룹핑
        addr_groups = {}
        for r in needs_geocoding:
            key = (r['주소'], r['도로명주소'])
            addr_groups.setdefault(key, []).append(r)
        print(f'고유 주소 그룹: {len(addr_groups):,}건', flush=True)

        resolved = {}
        counter = {'done': 0, 'exact': 0, 'approx': 0, 'fail': 0}
        lock = threading.Lock()
        total = len(addr_groups)

        def geocode_key(key):
            jibun, road = key
            acc, lat, lng = resolve_coords(jibun, road)
            return key, acc, lat, lng

        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            futures = [ex.submit(geocode_key, k) for k in addr_groups.keys()]
            for fut in as_completed(futures):
                key, acc, lat, lng = fut.result()
                resolved[key] = (acc, lat, lng)
                with lock:
                    counter['done'] += 1
                    if acc == 'exact':
                        counter['exact'] += 1
                    elif acc == 'approximate':
                        counter['approx'] += 1
                    else:
                        counter['fail'] += 1
                    if counter['done'] % 200 == 0 or counter['done'] == total:
                        print(
                            f"  [{counter['done']:,}/{total:,}] "
                            f"exact={counter['exact']:,} "
                            f"approx={counter['approx']:,} "
                            f"fail={counter['fail']:,}",
                            flush=True)

        # 레코드에 좌표 적용
        for r in needs_geocoding:
            key = (r['주소'], r['도로명주소'])
            acc, lat, lng = resolved[key]
            r['lat'] = lat
            r['lng'] = lng
            r['좌표정확도'] = acc

        print(f'\n지오코딩 결과: exact={counter["exact"]:,} approx={counter["approx"]:,} fail={counter["fail"]:,}', flush=True)
    else:
        counter = {'exact': 0, 'approx': 0, 'fail': 0}
        print('지오코딩 불필요 (전부 시드 재사용)', flush=True)

    # 5. 무결성 자체검증
    print('\n=== 무결성 검증 ===', flush=True)
    errors = []

    # 계기번호 중복
    meter_nos = [r['계기번호'] for r in new_records]
    dup = len(meter_nos) - len(set(meter_nos))
    if dup > 0:
        errors.append(f'계기번호 중복 {dup}건')
    else:
        print(f'  계기번호 중복: 0건 (OK)', flush=True)

    # 11자리 검증
    non11 = [r['계기번호'] for r in new_records if len(r['계기번호']) != 11]
    if non11:
        errors.append(f'11자리 아닌 계기번호 {len(non11)}건: {non11[:5]}')
    else:
        print(f'  계기번호 11자리: 전부 OK', flush=True)

    # 하이픈 검증
    hyphen = [r['계기번호'] for r in new_records if '-' in r['계기번호']]
    if hyphen:
        errors.append(f'하이픈 계기번호 {len(hyphen)}건')
    else:
        print(f'  하이픈: 0건 (OK)', flush=True)

    # 계기타입 vocab
    known_vocab = {'보안계기', 'AE타입', 'G타입', '알수없음', 'EA', 'E'}
    unknown_types = set(r['계기타입'] for r in new_records) - known_vocab
    if unknown_types:
        errors.append(f'미지 계기타입 vocab: {unknown_types}')
    else:
        print(f'  계기타입 vocab: {set(r["계기타입"] for r in new_records)} (OK)', flush=True)

    # '0' 값 검증 (변대주/상호/인입주)
    zero_vals = sum(
        1 for r in new_records
        if r.get('변대주') == '0' or r.get('상호') == '0' or r.get('인입주') == '0'
    )
    if zero_vals > 0:
        errors.append(f"'0' 값 잔존: {zero_vals}건")
    else:
        print(f"  '0' 값 잔존: 0건 (OK)", flush=True)

    # lat/lng null 비율
    null_coords = [r['계기번호'] for r in new_records if r['lat'] is None or r['lng'] is None]
    total_n = len(new_records)
    if null_coords:
        errors.append(f'lat/lng null: {len(null_coords)}건 ({len(null_coords)/total_n*100:.1f}%): {null_coords[:3]}')
    else:
        print(f'  lat/lng null: 0건 (OK)', flush=True)

    # 좌표 분포
    exact_n = sum(1 for r in new_records if r.get('좌표정확도') == 'exact')
    approx_n = sum(1 for r in new_records if r.get('좌표정확도') == 'approximate')
    fail_n = sum(1 for r in new_records if r.get('좌표정확도') == 'fail')
    seed_n = seeded
    print(f'  좌표 분포: 시드재사용={seed_n:,} | 지오코딩exact={counter["exact"]:,} approx={counter["approx"]:,} fail={counter["fail"]:,}', flush=True)
    print(f'  전체 exact={exact_n:,} approx={approx_n:,} fail={fail_n:,}', flush=True)

    if errors:
        print('\n[오류]', flush=True)
        for e in errors:
            print(f'  - {e}', flush=True)
        sys.exit(1)
    else:
        print('\n무결성 검증 통과', flush=True)

    # 6. 출력
    with open(OUT_JSON, 'w') as f:
        json.dump(new_records, f, ensure_ascii=False, indent=2)
    print(f'\n저장 완료: {OUT_JSON}', flush=True)
    print(f'신규 레코드 수: {len(new_records):,}건', flush=True)

    # 최종 보고
    print('\n=== 최종 보고 ===', flush=True)
    print(f'신규 레코드: {len(new_records):,}건', flush=True)
    print(f'계기타입 분포:', flush=True)
    for k, v in sorted(meter_type_dist.items(), key=lambda x: -x[1]):
        print(f'  {k}: {v:,}건', flush=True)
    print(f'좌표: exact={exact_n:,} approx={approx_n:,} fail={fail_n:,}', flush=True)
    print(f'  (좌표 시드 재사용: {seed_n:,}건)', flush=True)


if __name__ == '__main__':
    main()
