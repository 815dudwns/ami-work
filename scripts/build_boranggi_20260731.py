#!/usr/bin/env python3
"""주덕기 보강현황 20260730 → 아미맵 2분할 데이터셋 생성 (재 / 0731)

입력:
  - data/inbox/boranggi_20260730.xlsx        (활성행 10,679)
  - data/site-data.json                       (기존 미완료 원장)
  - data/site-data-completed-archive-20260704.json  (완료 아카이브)
  - data/rework-data.json                     (재방문 원장 1,805)
  - ~/Downloads/DCU_철거_예정_개소_목록.xlsx  (해지 164 / 유지 11)

출력 (data/inbox/ 아래, 원장 직접수정 금지):
  - boranggi-20260731-jae.json    재 리스트  (남은 재 1,695 + 새 재 7)
  - boranggi-20260731-new.json    0731 리스트 (이번 리스트에서 재 아닌 것)
  - boranggi-20260731-report.json 집계 리포트

규칙:
  - '재' = 완료 이력이 있는데 이번 리스트에 다시 나온 것 (재방문 원장 ∪ 완료 아카이브)
  - 변대주 필드에 DCU 철거예정 표기:
      해지 → '<변대주> / DCU 철거예정 개소 LTE 시설'
      유지 → '<변대주> / DCU 유지'
    매칭키 = 변대주명(공백제거 대문자). DCU 목록의 '변대주번호'는 보강의 DCU ID와
    체계가 달라 매칭 불가 — 이름 매칭만 유효하며 전 건 지사까지 일치함(오탐 0).
  - 좌표는 기존 원장에서 최대한 재사용, 나머지만 카카오→네이버→동중심 캐스케이드
"""
import json
import re
import sys
import collections
from pathlib import Path

from openpyxl import load_workbook

BASE = Path('/Users/woodelight/Projects/ami-work')
sys.path.insert(0, str(BASE / 'scripts'))

from build_boranggi_20260704 import (  # noqa: E402
    normalize_meter, normalize_cust_no, normalize_str, normalize_date,
    parse_meter_type, get_hidden_row_set, resolve_coords,
)

EXCEL = BASE / 'data/inbox/boranggi_20260730.xlsx'
DCU_XLSX = Path.home() / 'Downloads/DCU_철거_예정_개소_목록.xlsx'
OUT_DIR = BASE / 'data/inbox'

# 보강 엑셀 열 인덱스 (0-base)
C_DEPT, C_FAULT, C_SYSREG, C_EXPIRE = 0, 1, 2, 3
C_SWAP, C_LINK, C_LP = 4, 5, 6
C_DCUID, C_MAC, C_METER = 10, 11, 12
C_TYPE, C_CHA, C_COMM, C_BDJ = 16, 17, 18, 19
C_JIBUN, C_ROAD, C_APT, C_BIZ = 20, 21, 22, 23


def nkey(s):
    return re.sub(r'\s+', '', str(s or '')).upper()


def load_ledger(path):
    with open(path) as f:
        d = json.load(f)
    return d if isinstance(d, list) else (d.get('data') or d.get('sites') or [])


def main():
    print('보강 엑셀 파싱...', flush=True)
    hidden = get_hidden_row_set(EXCEL)
    ws = load_workbook(EXCEL, data_only=True, read_only=True).active
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if i in hidden:
            continue
        mid = normalize_meter(r[C_METER])
        if not mid or mid in ('#N/A', 'None'):
            continue
        rows.append({'idx': i, 'r': r, 'mid': mid})
    print(f'활성행 {len(rows):,}', flush=True)

    # ── 원장 로드 ────────────────────────────────────────────────
    todo = load_ledger(BASE / 'data/site-data.json')
    done = load_ledger(BASE / 'data/site-data-completed-archive-20260704.json')
    rew = load_ledger(BASE / 'data/rework-data.json')
    done_mid = {normalize_meter(x.get('계기번호')) for x in done}
    rew_mid = {normalize_meter(x.get('계기번호')) for x in rew}
    print(f'원장 — 미완료 {len(todo):,} / 완료 {len(done):,} / 재방문 {len(rew):,}', flush=True)

    # 기존 데이터에서 값을 끌어올 때의 키 = 고객번호 > 계기번호+주소 (scripts/ledger_match.py).
    #   ★계기번호 단독 매칭 금지(영준님 2026-08-03). 검기만료가 아닌 사유로 철거 후 재시공하면
    #     같은 계기번호가 다른 개소에 나타나 엉뚱한 인입주·좌표가 조용히 붙는다.
    from ledger_match import build_index, lookup  # noqa: E402
    _by_cust, _by_pair = build_index(todo + done + rew)

    # 좌표 폴백용 주소 시드 — 좌표는 개소가 아니라 '주소'에 붙는 값이라 주소 일치면 안전하다.
    seed_addr = {}
    for x in todo + done + rew:
        if x.get('lat') is None:
            continue
        a = normalize_str(x.get('주소'))
        if a:
            seed_addr.setdefault(a, x)

    # ── DCU 철거예정 목록 ────────────────────────────────────────
    dws = load_workbook(DCU_XLSX, data_only=True).worksheets[0]
    dcu_by_name = {}
    for r in dws.iter_rows(min_row=4, values_only=True):
        if not r[4]:
            continue
        dcu_by_name.setdefault(nkey(r[5]), []).append({
            'dept': normalize_str(r[1]), 'verdict': normalize_str(r[7]),
        })
    print(f'DCU 변대주 {len(dcu_by_name):,}개', flush=True)

    # ── 레코드 생성 ──────────────────────────────────────────────
    need_geo = []
    recs = []
    for row in rows:
        r, mid = row['r'], row['mid']
        jibun = normalize_str(r[C_JIBUN])
        road = normalize_str(r[C_ROAD])
        bdj = normalize_str(r[C_BDJ])

        # DCU 표기
        dcu_tag = ''
        hit = dcu_by_name.get(nkey(bdj))
        if hit:
            dept = normalize_str(r[C_DEPT])
            m = next((h for h in hit if h['dept'] == dept), None)
            if m:
                dcu_tag = ('DCU 철거예정 개소 LTE 시설' if m['verdict'] == '해지'
                           else 'DCU 유지')
        # ★변대주 값에는 태그를 섞지 않는다(영준님 2026-08-02). 태그는 dcu_철거예정
        #   필드로만 싣고, 표시는 js/detail.js 의 큰 글씨 공통줄이 담당한다.
        bdj_out = bdj

        rec = {
            '지사': normalize_str(r[C_DEPT]),
            '주소': jibun,
            '도로명주소': road,
            '계기번호': mid,
            '계기타입': parse_meter_type(mid),
            '고객번호': '',
            '통신방식': normalize_str(r[C_COMM]),
            '공동주택명': normalize_str(r[C_APT]),
            '상호': normalize_str(r[C_BIZ]),
            '검기만료년월': normalize_str(r[C_EXPIRE]),
            '계기타입_전': normalize_str(r[C_TYPE]),
            '인입주': bdj,
            '변대주': bdj_out,
            'DCUID': normalize_str(r[C_DCUID]),
            '모뎀MAC': normalize_str(r[C_MAC]),
            'lat': None, 'lng': None, '좌표정확도': '',
            'DCU장애여부': normalize_str(r[C_FAULT]),
            '교체사유': '보강',
            '시스템등록일': normalize_date(r[C_SYSREG]),
            '계기교체일': normalize_date(r[C_SWAP]),
            '연계수신일': normalize_date(r[C_LINK]),
            '최초LP수신일': normalize_date(r[C_LP]),
            '사업차수_전': normalize_str(r[C_CHA]),
            'dcu_철거예정': dcu_tag,
        }

        # 기존 데이터에서 끌어오기 — 고객번호 > 계기번호+주소. 못 찾으면 비운다.
        #   ※ 주덕기 엑셀에는 고객번호 열이 없어 현재는 사실상 2순위만 걸린다.
        #     고객번호 열을 받게 되면 자동으로 1순위가 쓰인다.
        s, _how = lookup(rec, _by_cust, _by_pair)
        if s:
            rec['고객번호'] = normalize_cust_no(s.get('고객번호')) or ''
            # 인입주는 이 엑셀에 열 자체가 없다. 변대주를 복사하면 '두 기둥이 같다'는
            # 거짓이 되므로, 기존 데이터에 진짜 값이 있을 때만 싣는다.
            rec['인입주'] = normalize_str(s.get('인입주'))
            rec['match_by'] = _how

        # 좌표 — 개소가 아니라 주소에 붙는 값이라 주소 일치면 안전하다.
        c = s if (s and s.get('lat') is not None) else seed_addr.get(jibun)
        if c and c.get('lat') is not None:
            rec['lat'], rec['lng'] = c.get('lat'), c.get('lng')
            rec['좌표정확도'] = c.get('좌표정확도') or ''
        else:
            need_geo.append(rec)

        # 재 판정
        rec['_jae'] = mid in done_mid or mid in rew_mid
        recs.append(rec)

    print(f'좌표 재사용 {len(recs) - len(need_geo):,} / 지오코딩 필요 {len(need_geo):,}', flush=True)

    # ── 지오코딩 (주소 단위 캐시) ────────────────────────────────
    cache = {}
    for i, rec in enumerate(need_geo, 1):
        k = (rec['주소'], rec['도로명주소'])
        if k not in cache:
            cache[k] = resolve_coords(rec['주소'], rec['도로명주소'])
        acc, lat, lng = cache[k]
        rec['좌표정확도'], rec['lat'], rec['lng'] = acc, lat, lng
        if i % 200 == 0:
            print(f'  지오코딩 {i:,}/{len(need_geo):,} (고유주소 {len(cache):,})', flush=True)
    fails = [r for r in need_geo if r['lat'] is None]
    print(f'지오코딩 완료 — 고유주소 {len(cache):,}, 실패 {len(fails):,}', flush=True)

    # ── 2분할 ────────────────────────────────────────────────────
    jae_new = [r for r in recs if r['_jae']]          # 이번 리스트에 나온 재
    lst_new = [r for r in recs if not r['_jae']]      # 0731 신규 리스트
    jae_mid = {r['계기번호'] for r in jae_new}

    # 이번 리스트에 없는 '남은 재'(재방문 원장 미완료분)도 재 리스트에 살림.
    #   ※ 재방문 원장 1,805 중 이미 complete 처리된 건은 제외 — 라이브 workStatus 기준.
    ws_live = json.loads(
        (BASE / 'data/ws-live-snapshot.json').read_text()) if (BASE / 'data/ws-live-snapshot.json').exists() else {}

    def is_done(x):
        e = ws_live.get(normalize_str(x.get('도로명주소'))) or ws_live.get(normalize_str(x.get('주소')))
        return bool(e) and e.get('state') == 'complete'

    carry = [x for x in rew
             if normalize_meter(x.get('계기번호')) not in jae_mid and not is_done(x)]

    for r in recs:
        r.pop('_jae', None)
    for r in jae_new:
        r['category'] = '재방문'
        r['rework'] = True
    for r in lst_new:
        r['category'] = '실효'
    for x in carry:
        x['category'] = '재방문'
        x['rework'] = True

    jae_all = jae_new + carry

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_DIR / 'boranggi-20260731-jae.json', 'w') as f:
        json.dump(jae_all, f, ensure_ascii=False, indent=1)
    with open(OUT_DIR / 'boranggi-20260731-new.json', 'w') as f:
        json.dump(lst_new, f, ensure_ascii=False, indent=1)

    report = {
        '활성행': len(rows),
        '재_이번리스트': len(jae_new),
        '재_이월(리스트에없는 남은재)': len(carry),
        '재_합계': len(jae_all),
        '0731_신규리스트': len(lst_new),
        '좌표_지오코딩': len(need_geo),
        '좌표_실패': len(fails),
        'DCU_해지표기': sum(1 for r in recs if r['dcu_철거예정'].startswith('DCU 철거예정')),
        'DCU_유지표기': sum(1 for r in recs if r['dcu_철거예정'] == 'DCU 유지'),
        '지사별_0731': dict(collections.Counter(r['지사'] for r in lst_new)),
        '타입별_0731': dict(collections.Counter(r['계기타입'] for r in lst_new)),
        '정확도_0731': dict(collections.Counter(r['좌표정확도'] for r in lst_new)),
    }
    with open(OUT_DIR / 'boranggi-20260731-report.json', 'w') as f:
        json.dump(report, f, ensure_ascii=False, indent=1)
    print(json.dumps(report, ensure_ascii=False, indent=1))


if __name__ == '__main__':
    main()
