import openpyxl, glob, json, collections, re, sys
sys.path.insert(0,'/Users/woodelight/Projects/ami-work/scripts')
from concurrent.futures import ThreadPoolExecutor
from geocode_cascade import resolve
SC='/private/tmp/claude-501/-Users-woodelight-Projects-ami-work/eeaa7f75-b1bf-487b-ac1d-efa6c00c873a/scratchpad'
def mno(s):
    s=str(s or '').strip().replace('-',''); return s.zfill(11) if s.isdigit() else s
# 원본 엑셀 찾기 — 경로가 아니라 **시트 구성**으로 고른다.
#   파일명이 압축해제에서 깨지기도 하고(cp949), 스크래치패드는 주기적으로 비워진다.
#   인자로 경로를 주면 그걸 쓴다: python3 scripts/build_jangae_data.py <xlsx>
def _find_src():
    if len(sys.argv) > 1:
        return sys.argv[1]
    cands = [p for p in glob.glob(SC + '/**/*.xlsx', recursive=True) if '~$' not in p]
    cands += [p for p in glob.glob('data/**/*.xlsx', recursive=True) if '~$' not in p]
    for p in sorted(cands, key=lambda x: -len(x)):
        try:
            wb0 = openpyxl.load_workbook(p, read_only=True)
            names = set(wb0.sheetnames); wb0.close()
        except Exception:
            continue
        if {'장애', '모뎀작업리스트'} <= names:
            return p
    sys.exit('모뎀작업리스트 엑셀을 못 찾았다 — 경로를 인자로 넘겨라')

f=_find_src()
print('원본:', f, flush=True)
wb=openpyxl.load_workbook(f, data_only=True, read_only=True)
def sh(nm):
    ws=wb[nm]; it=ws.iter_rows(values_only=True)
    h=[str(x).strip() if x is not None else '' for x in next(it)]
    return h,[r for r in it if any(x is not None and str(x).strip()!='' for x in r)]
h1,ja=sh('장애'); h2,full=sh('모뎀작업리스트'); wb.close()
I1={k:i for i,k in enumerate(h1)}; I2={k:i for i,k in enumerate(h2)}
def C(r,I,k):
    i=I.get(k); x=r[i] if i is not None and i<len(r) else None
    if x is None: return ''
    if isinstance(x,float) and x.is_integer(): x=int(x)
    s=str(x).strip(); return '' if s in ('None','nan','-') else s

# 장애 시트: MAC -> 주소, 장애계기 집합
# ★주소는 그룹 안에서 다수결로 정한다(영준님 2026-09-02).
#   예전엔 setdefault 로 '첫 행'을 썼는데, 첫 행이 마스터라 마스터 주소가 오입력이면
#   그 소수 주소가 그룹 전체를 대표해버렸다. 실제로 01253654870(27계기)이 슬레이브 26건은
#   성북구 보문동7가인데 마스터 1건만 광진구 화양동이라 지도에 광진구로 찍혔다.
#   한 모뎀에 물린 계기는 한 함체 안이다 — 갈리면 많은 쪽이 맞다.
#   동수면 마스터 주소를 쓴다(그 이상 가릴 근거가 없다).
_addr_votes=collections.defaultdict(collections.Counter)
_addr_master={}
FAILSET=collections.defaultdict(set)
for r in ja:
    m=C(r,I1,'기존모뎀MAC'); a=C(r,I1,'주소')
    if a: _addr_votes[m][a]+=1
    if C(r,I1,'모뎀유형')=='마스터' and a: _addr_master.setdefault(m, a)
    FAILSET[m].add(mno(C(r,I1,'계기번호')))
# ★번지가 있는 주소를 먼저 본다(영준님 2026-09-04 "주소에는 망우동밖에 안나왓네").
#   한전 원본은 같은 함체인데도 어떤 행은 '중랑구 망우동' 까지만, 어떤 행은 '망우동 419-23'
#   으로 들어온다. 표수만 세면 번지 없는 쪽이 이기고(실측 01254423544: 망우동 7 vs 419-23 2),
#   지오코딩이 동 중심을 돌려줘 마커가 실제 위치에서 470m 떨어진 곳에 박힌다.
#   번지 없는 주소는 '덜 적힌 것'이지 다른 주소가 아니다 — 정보가 더 있는 쪽이 맞다.
#   번지 있는 후보들 사이에서만 다수결하고, 하나도 없으면 종전대로 전체 다수결.
_BEONJI_RE = re.compile(r'(?:동|가|읍|면)\d*\s+산?\s*\d+(?:-\d+)?')
def _has_beonji(a): return bool(_BEONJI_RE.search(str(a or '')))

mac_addr={}
for m, votes in _addr_votes.items():
    pool = {a:n for a,n in votes.items() if _has_beonji(a)} or dict(votes)
    top=sorted(pool.items(), key=lambda kv:(-kv[1], kv[0]))
    best_n=top[0][1]
    tied=[a for a,n in top if n==best_n]
    mac_addr[m] = _addr_master[m] if (len(tied)>1 and _addr_master.get(m) in tied) else tied[0]
    if len(votes)>1:
        print(f'  [주소 다수결] {m}: {mac_addr[m]} ({votes[mac_addr[m]]}건) '
              f'<- 후보 {dict(votes)}', flush=True)
for m in _addr_votes:
    pass
# 주소가 아예 없는 MAC 도 그룹으로는 남긴다(좌표 실패로 처리)
for r in ja:
    mac_addr.setdefault(C(r,I1,'기존모뎀MAC'), '')
# 시트2: MAC -> 전체 행
F=collections.defaultdict(list)
for r in full: F[C(r,I2,'기존모뎀MAC')].append(r)
print(f'장애 MAC {len(mac_addr)} / 장애계기 {sum(len(v) for v in FAILSET.values())}', flush=True)
tot=sum(len(F.get(m,[])) for m in mac_addr)
print(f'시트2 기준 그룹 전체 계기 {tot}', flush=True)

# 진짜 DCUID·변대주명 = 우리 데이터에서 계기번호 매칭
REAL={}
for p in ['data/site-data.json','data/rework-data.json','data/hapdong-data.json',
          'data/hapdong-data-archive.json','data/site-data-completed-archive-20260704.json',
          'data/site-data.backup-리스트업전-20260831-030051.json']:
    try: d=json.load(open(p))
    except Exception: continue
    for x in d:
        m=mno(x.get('계기번호')); du=str(x.get('DCUID') or '').strip(); bj=str(x.get('변대주') or '').strip()
        if m and (du or bj): REAL.setdefault(m,(du,bj))

addrs=sorted(set(mac_addr.values()))
def work(a):
    try: return a, resolve(jibun=a, road='')
    except Exception: return a, None
geo={}; n=0
with ThreadPoolExecutor(max_workers=8) as ex:
    for a,hit in ex.map(work, addrs):
        geo[a]=hit; n+=1
        if n%80==0: print(f'  geo {n}/{len(addrs)}', flush=True)
print('좌표:', dict(collections.Counter(h.accuracy if h else 'fail' for h in geo.values())), flush=True)

out=[]
for mac, addr in mac_addr.items():
    rs=F.get(mac) or [r for r in ja if C(r,I1,'기존모뎀MAC')==mac]
    src2 = mac in F
    rep=next((r for r in rs if C(r,I2 if src2 else I1,'모뎀유형')=='마스터'), rs[0])
    II = I2 if src2 else I1
    rd=rb=''
    for r in rs:
        v=REAL.get(mno(C(r,II,'계기번호')))
        if v:
            if not rd and v[0]: rd=v[0]
            if not rb and v[1]: rb=v[1]
        if rd and rb: break
    meters=[]
    for r in rs:
        m=mno(C(r,II,'계기번호'))
        meters.append({
            '계기번호': m, '계기타입': C(r,II,'계기유형'), '모뎀유형': C(r,II,'모뎀유형'),
            '시설유형': C(r,II,'시설유형'), '작업구분': C(r,II,'작업구분'),
            '분기기': C(r,II,'분기기'), 'LP': C(r,II,'LP'),
            '상태': C(r,II,'상태') if src2 else '',
            '장애': m in FAILSET[mac],
        })
    hit=geo.get(addr)
    out.append({
        '지사': C(rep,II,'지사'), '주소': addr,
        '도로명주소': (hit.road if hit and hit.accuracy!='fail' and hit.road else ''),
        'lat': hit.lat if hit and hit.accuracy!='fail' else None,
        'lng': hit.lng if hit and hit.accuracy!='fail' else None,
        '좌표정확도': hit.accuracy if hit else 'fail',
        '모뎀MAC': mac, '기술타입': C(rep,II,'기술타입'),
        '변대주': C(rep,II,'변대주번호'), 'DCUID': rd, '변대주명': rb,
        '작업일': C(rep,II,'작업일자'), '작업자1': C(rep,II,'작업자1'), '작업자2': C(rep,II,'작업자2'),
        '외장형연결장치': C(rep,II,'외장형연결장치'), '개통여부': C(rep,II,'개통여부'),
        '진행상태': C(rep,II,'진행 상태'), '사업번호': C(rep,II,'사업번호'),
        '계기수': len(meters), '장애수': sum(1 for m in meters if m['장애']),
        '실패수': sum(1 for m in meters if m['상태']=='실패'),
        '계기목록': meters,
    })
print(f'\n그룹 {len(out)}개 / 계기 {sum(x["계기수"] for x in out)} (장애 {sum(x["장애수"] for x in out)})')
print('  시트2에서 보강된 계기:', sum(x['계기수'] for x in out)-sum(x['장애수'] for x in out))
print('  DCUID 채움:', sum(1 for x in out if x['DCUID']), '/', len(out))
print('  개통:', dict(collections.Counter(x['개통여부'] or '(빈칸)' for x in out)))
print('  장애수 분포:', dict(sorted(collections.Counter(x['장애수'] for x in out).items())[:12]))
json.dump(out, open('data/jangae-data.json','w'), ensure_ascii=False)
print('저장: data/jangae-data.json')
