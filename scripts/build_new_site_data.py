#!/usr/bin/env python3
"""
새 site-data.json 생성:
- 새 DB Sheet1 (작업대상 16,490건) 베이스
- 기존 site-data와 겹치는 건 → 기존 변대주(DCUID) + 기존 좌표 유지
- 신규 건 → 보강현황 메인의 한글 변대주 + 좌표 None (별도 추출)
- 빠진 건 → 제거
"""

import json
import os
from openpyxl import load_workbook
from collections import Counter

BASE = '/Users/woodelight/Projects/ami-work'
NEW_XLSX = '/Users/woodelight/Library/Mobile Documents/com~apple~CloudDocs/계기교체 보강현황_20260513155458.xlsx'
OLD_JSON = os.path.join(BASE, 'data/site-data.json')
OUT_JSON = os.path.join(BASE, 'data/site-data-new.json')
COORDS_TODO = os.path.join(BASE, 'data/coords-todo.json')

def map_type(t):
    if not t: return '알수없음'
    if '보안' in t: return 'Amigo'
    if 'AE' in t.upper(): return 'EA'
    if 'G타입' in t: return 'G'
    if 'E타입' in t: return 'E'
    return t

def main():
    # 1. 보강현황 메인 시트 → 계기번호 인덱스 (보강 후 정보 위주)
    wb = load_workbook(NEW_XLSX, read_only=True, data_only=True)
    s1_main = {}
    for r in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True):
        if r[16]:
            m = str(r[16]).strip().zfill(11)
            s1_main[m] = r
    
    # 2. Sheet1 (작업대상) 로드
    target_rows = []
    for r in wb['Sheet1'].iter_rows(min_row=2, values_only=True):
        if r[3]:
            target_rows.append(r)
    wb.close()
    
    print(f"새 작업대상: {len(target_rows)}건")
    print(f"보강현황 메인: {len(s1_main)}건")
    
    # 3. 기존 DB 인덱스
    with open(OLD_JSON) as f:
        old = json.load(f)
    old_by_meter = {str(d['계기번호']).strip().zfill(11): d for d in old}
    print(f"기존 DB: {len(old_by_meter)}건")
    
    # 4. 가공
    new_entries = []
    coords_todo = []
    
    stats = {
        '겹침_기존변대주살림': 0,
        '신규_새변대주': 0,
        '신규_변대주없음': 0,
    }
    
    for s2 in target_rows:
        m = str(s2[3]).strip().zfill(11)
        s1 = s1_main.get(m)
        old_e = old_by_meter.get(m)
        
        entry = {
            '지사': s2[0],
            '주소': (s2[5] or '').strip(),
            '도로명주소': (s2[6] or '').strip(),
            '계기번호': m,
            '계기타입': map_type(s2[4]),
            '고객번호': str(s2[1]).zfill(10) if s2[1] else '',
            '통신방식': s2[2] or '',
            '공동주택명': s2[7] or '',
            '상호': (s1[31] if s1 and s1[31] else ''),
            '검기만료년월': (s1[5] if s1 and s1[5] else ''),
            '계기타입_전': map_type(s1[10]) if s1 else '',
            '인입주': (s1[27] if s1 and s1[27] else ''),
        }
        
        # 변대주/DCUID/좌표 결정
        if old_e:
            # 겹침 → 기존 살림
            entry['변대주'] = old_e.get('변대주', '')
            entry['DCUID'] = old_e.get('DCUID', '')
            entry['lat'] = old_e.get('lat')
            entry['lng'] = old_e.get('lng')
            entry['좌표정확도'] = old_e.get('좌표정확도', 'exact')
            stats['겹침_기존변대주살림'] += 1
        else:
            # 신규 → 보강현황 메인의 한글 변대주만
            변대주 = (s1[26] if s1 and s1[26] else '')
            entry['변대주'] = 변대주
            entry['DCUID'] = ''  # 보강 후 DCU ID 없음
            entry['lat'] = None
            entry['lng'] = None
            entry['좌표정확도'] = None
            if 변대주:
                stats['신규_새변대주'] += 1
            else:
                stats['신규_변대주없음'] += 1
            coords_todo.append({
                '계기번호': m,
                '주소': entry['주소'],
                '도로명주소': entry['도로명주소'],
            })
        
        new_entries.append(entry)
    
    print(f"\n=== 가공 통계 ===")
    for k, v in stats.items():
        print(f"  {k}: {v}건")
    print(f"  총 출력: {len(new_entries)}건")
    print(f"  좌표 추출 대기: {len(coords_todo)}건")
    
    # 통신방식 분포
    print(f"\n=== 통신방식 분포 ===")
    for k, v in Counter(e['통신방식'] for e in new_entries).most_common():
        print(f"  {k}: {v}")
    
    # 계기타입 분포
    print(f"\n=== 계기타입 분포 ===")
    for k, v in Counter(e['계기타입'] for e in new_entries).most_common():
        print(f"  {k}: {v}")
    
    with open(OUT_JSON, 'w') as f:
        json.dump(new_entries, f, ensure_ascii=False, indent=2, default=str)
    with open(COORDS_TODO, 'w') as f:
        json.dump(coords_todo, f, ensure_ascii=False, indent=2)
    
    print(f"\n저장: {OUT_JSON}")
    print(f"저장: {COORDS_TODO}")

if __name__ == '__main__':
    main()
