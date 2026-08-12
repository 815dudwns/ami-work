#!/usr/bin/env python3
"""한전 DCU 자료를 site-data 형식 데이터셋에 붙인다 (재사용 도구).

새 리스트가 들어올 때마다 반드시 한 번 돌린다 — 데이터셋 추가 프로세스의 고정 단계.

사용:
    python3 scripts/apply_dcu_status.py data/site-data.json data/rework-data.json
    python3 scripts/apply_dcu_status.py --dry-run data/site-data.json

붙이는 값 두 가지:

  1) dcu_철거예정 — '검토' 시트(철거 예정 개소)
        해지 -> 'DCU 철거예정 개소 LTE 시설'
        유지 -> 'DCU 유지'
     함께 dcu_잔여호수 / dcu_전체호수 도 싣는다(2026-08-06 자료부터 제공).

  2) 통신방식 — '전체DCU 현황' 시트(서울 전체 DCU 대장 19,007행)
     PLC / HPGP / K-DCU. **한전 원본 값이라 추정이 아니다.**
     주덕기 보강 엑셀은 통신방식이 1%밖에 안 차 있어 이쪽에서 채운다.
     이미 값이 있으면 덮지 않는다.

★매칭 규칙 (실측으로 확정)
  - 1순위 변대주명 + 지사. 지사까지 대조해야 동명 변대주 오탐이 없다.
  - 2순위 DCU ID 또는 변대주번호.
  - DCU 목록의 '변대주번호'(0028G232)와 보강 엑셀 'DCU ID'(9726G1525M)는 체계가 달라
    번호 단독 매칭은 신뢰도가 낮다. 그래서 이름을 먼저 본다.
  - 태그·통신방식은 각자 필드에만 싣는다. 변대주 값에 문자열로 섞지 마라 —
    표시는 js/detail.js 가 담당한다.

자료 정본: data/reference/간선망_해지_정지대상.xlsx
  (한전에서 새로 오면 이 파일을 교체하고 다시 돌린다)
"""
import json
import re
import sys
import collections
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
XLSX = BASE / 'data/reference/간선망_해지_정지대상.xlsx'

TAG_REMOVE = 'DCU 철거예정 개소 LTE 시설'
TAG_KEEP = 'DCU 유지'


def nkey(s):
    return re.sub(r'\s+', '', str(s or '')).upper()


def _header_row(ws, label='지사'):
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if label in [str(v or '').strip() for v in r]:
            return i
    return 2


def load_reference(path=XLSX):
    """(철거예정 인덱스, 통신방식 인덱스) 반환. 각각 (이름키, 번호키) 쌍."""
    wb = load_workbook(path, data_only=True, read_only=True)

    # --- 검토 시트: 철거 예정 개소 ---
    ws = wb['검토']
    start = _header_row(ws) + 1
    rm_name, rm_id = collections.defaultdict(list), {}
    for r in ws.iter_rows(min_row=start, values_only=True):
        if not r[3]:
            continue
        rec = {
            'dept': str(r[1] or '').strip(),
            'verdict': str(r[12] or '').strip(),
            'total': r[9], 'lte': r[10], 'rest': r[11],
        }
        rm_name[(rec['dept'], nkey(r[7]))].append(rec)
        for k in (nkey(r[3]), nkey(r[6])):
            if k:
                rm_id.setdefault(k, rec)

    # --- 전체DCU 현황: 인입망 통신방식 ---
    ws2 = wb['전체DCU 현황']
    start2 = _header_row(ws2) + 1
    cm_name, cm_id = collections.defaultdict(collections.Counter), {}
    # 변대주명 -> 변대주번호. 대장이 이름과 번호를 같이 들고 있어 우리 데이터에 없는 번호를
    # 채울 수 있다(영준님 2026-08-06). 한 이름에 번호가 여럿이면 애매하므로 싣지 않는다.
    pn_name = collections.defaultdict(set)
    for r in ws2.iter_rows(min_row=start2, values_only=True):
        if not r[3]:
            continue
        dept, nm = str(r[1] or '').strip(), nkey(r[7])
        pole_no = str(r[6] or '').strip()
        if pole_no:
            pn_name[(dept, nm)].add(pole_no)
        comm = str(r[8] or '').strip()
        if not comm:
            continue
        # ★회선이 죽은 DCU 는 통신방식을 싣지 않는다 (영준님 2026-08-12: "안 써있어야지").
        #   이 시트의 '인입망 통신방식' 칸은 19,007행 전부 채워져 있다 — DCU 가 등록된 방식일 뿐
        #   지금 살아 있다는 뜻이 아니다. 생사는 '회선상태'(r[12], 개통/정지/해지)와
        #   '장애여부'(r[3])가 말한다. 이 조건이 없어서 해지된 DCU 의 PLC 가 개소에 붙었다
        #   (실측: 동명간 29L6R1 = 회선 해지·계기 0/0 인데 23개 개소에 PLC).
        #   ★'계기 없음'은 제외 조건이 아니다 — DCU 는 살아 있고 아직 계기가 안 물린 것뿐이라
        #     오히려 우리가 PLC 로 시공할 대상이다(2026-08-12 실측 중 과잉 제거 127건을 되돌렸다).
        #   ★인덱스 주의: r[12]=통신사, r[13]=회선상태. 처음에 r[12]로 잘못 짚어 가드가
        #     통째로 무력화됐고(통신사 값은 해지·정지가 아니므로 전부 통과) 죽은 회선에
        #     PLC 가 96건 다시 붙었다. 컬럼을 바꿀 땐 헤더를 실제로 찍어 확인하라.
        line_state = str(r[13] or '').strip()
        if line_state in ('해지', '정지'):
            continue
        cm_name[(dept, nm)][comm] += 1
        for k in (nkey(r[3]), nkey(r[6])):
            if k:
                cm_id.setdefault(k, comm)

    return (rm_name, rm_id), (cm_name, cm_id, pn_name)


def apply_to(records, removal, comm):
    rm_name, rm_id = removal
    cm_name, cm_id, pn_name = comm
    stat = collections.Counter()

    for x in records:
        # ★변대주 먼저. DCU 가 달리는 전주는 변대주이고, 인입주는 계기로 들어가는 다른 전주라
        #   대장의 '변대주명'과 맞출 값이 아니다(영준님 2026-08-12: "변대주가 중요하지 인입주는 아님").
        #   인입주를 앞에 두는 바람에 변대주로는 걸리는 개소 7건이 철거예정 태그를 못 받았다
        #   (그중 4건이 '해지' 판정 — 철거될 DCU 인데 표시가 없어 PLC 로 시공할 뻔했다).
        #   인입주는 변대주가 비었을 때의 폴백으로만 남긴다.
        name = (x.get('변대주') or x.get('인입주') or '').strip()
        dept = (x.get('지사') or '').strip()
        nk = nkey(name)
        did = nkey(x.get('DCUID'))

        # 1) 철거 예정
        hit = None
        for h in rm_name.get((dept, nk), []):
            if h['dept'] == dept:
                hit = h
                break
        if hit is None:
            hit = rm_id.get(did)
        tag = ''
        if hit:
            tag = TAG_REMOVE if hit['verdict'] == '해지' else TAG_KEEP
            x['dcu_전체호수'] = hit['total']
            x['dcu_잔여호수'] = hit['rest']
        x['dcu_철거예정'] = tag
        stat[tag or '(철거예정 없음)'] += 1

        # 2) 변대주번호는 싣지 않는다(영준님 2026-08-06). DCU ID 줄이 이미 10자리를 보여주고
        #    복사만 8자리로 자른다. 대장 변대주번호는 같을 땐 그 복사값과 중복이고, 실측
        #    6,779건 중 1,080건(16%)이 우리 DCUID와 아예 달랐다(동명 변대주로 추정).
        #    pn_name 인덱스는 대조·감사용으로 남겨둔다.

        # 3) 통신방식 — 이미 있으면 덮지 않는다
        if str(x.get('통신방식') or '').strip():
            stat['통신방식 원본유지'] += 1
            continue
        cand = cm_name.get((dept, nk))
        if cand:
            x['통신방식'] = cand.most_common(1)[0][0]
            x['통신방식_출처'] = '전체DCU(변대주명)'
            stat['통신방식 변대주명'] += 1
        elif cm_id.get(did):
            x['통신방식'] = cm_id[did]
            x['통신방식_출처'] = '전체DCU(DCU ID)'
            stat['통신방식 DCU ID'] += 1
        else:
            stat['통신방식 못채움'] += 1
    return stat


def main(argv):
    dry = '--dry-run' in argv
    targets = [a for a in argv if not a.startswith('--')]
    if not targets:
        print(__doc__)
        return 1

    removal, comm = load_reference()
    print(f'자료 {XLSX.name} — 철거예정 {len(removal[0]):,}개 변대주 / '
          f'전체DCU {len(comm[0]):,}개 변대주')

    for t in targets:
        p = Path(t)
        data = json.loads(p.read_text())
        recs = data if isinstance(data, list) else (data.get('data') or [])
        st = apply_to(recs, removal, comm)
        print(f'\n{"[DRY] " if dry else "[적용] "}{p.name}  총 {len(recs):,}')
        for k, v in st.most_common():
            print(f'    {k:22} {v:6,}')
        if not dry:
            p.write_text(json.dumps(data, ensure_ascii=False))

    if not dry:
        print('\n※ site-data.json 을 바꿨으면 scripts/gen_site_version.py 를 반드시 다시 돌려라.')
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
