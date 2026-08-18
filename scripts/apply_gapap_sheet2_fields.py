#!/usr/bin/env python3
"""고압철거 데이터를 원천 엑셀(Sheet2) 기준으로 맞춘다.
  1) 한전기준·비고1·비고2·주소2 부착
  2) 작업대상만 남기기(251 -> 168)
  3) 계기번호를 계기번호2(현재 실제 번호)로 교체하고 계기타입 재판정

사용:
    python3 scripts/apply_gapap_sheet2_fields.py                # 예행(dry-run) — 파일 안 씀
    python3 scripts/apply_gapap_sheet2_fields.py --apply        # 실제 반영(+백업)
    python3 scripts/apply_gapap_sheet2_fields.py --no-filter    # 필드만 붙이고 건수는 유지
    python3 scripts/apply_gapap_sheet2_fields.py --no-renumber  # 계기번호 교체 안 함

배경(2026-08-18 영준님 확정):
  - 원천 = data/inbox_20260813/25년보강_고압철거대상_v2.xlsx 의 **Sheet2만**.
    Sheet1 에는 '한전기준' 열이 아예 없다.
  - Sheet2 는 자동필터로 184행이 숨겨져 있지만 openpyxl 은 숨김과 무관하게 전부 읽는다.
  - 매칭키 = 계기번호 zfill(11). Sheet2 429행 계기번호 중복 0, gapap 251건 전부 매칭.

작업대상 필터(3조건 전부 만족, 2026-08-18 영준님 확정):
  1. 한전기준 이 있을 것 — 빈칸은 대상이 아니다(빠지는 83건이 전부 이것)
  2. 주소 가 있을 것 ('#N/A' 제외)
  3. 대상 == '철거'
  결과 168건 = 한전기준 '철거+재설치' 158 + '철거' 10, 고유 주소 146.
  ※실측상 251건은 전부 대상='철거'·주소 있음이라, 실제로 거르는 건 조건 1뿐이다.
    나머지 두 조건은 원천이 바뀌었을 때를 위한 방어다.

계기번호2 기준(2026-08-18 영준님 확정):
  "계기교체 등으로 번호가 바뀐 게 있어 계기번호2 기준으로 리스트를 꾸린다.
   계기번호1은 디테일에도 없앤다." -> 레코드의 `계기번호` 를 계기번호2 로 덮어쓰고,
  계기번호1 은 데이터에 남기지 않는다(추적은 gapap_계기번호2교체내역_*.json 파일로).
  계기타입은 새 번호 3~4번째 자리로 다시 판정한다(CLAUDE.md 데이터규칙).
  코드가 매핑에 없으면(38·52 등) 지어내지 않고 엑셀 Sheet2 계기타입을 그대로 쓴다.
  조회표는 계기번호1·계기번호2 양쪽으로 걸어 두어 이미 교체된 파일에 재실행해도 매칭된다.

정규화: 셀이 '0' / '#N/A' / 'nan' / None 이면 빈 문자열.

주의: 기존 `비고` 필드는 비고1+비고2 를 합쳐 담고 있던 것이라 **건드리지 않는다**(하위호환).
"""

import argparse
import collections
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TARGET = ROOT / "data" / "gapap-data.json"
XLSX = ROOT / "data" / "inbox_20260813" / "25년보강_고압철거대상_v2.xlsx"
SHEET = "Sheet2"

# 엑셀 열 이름 -> gapap-data.json 필드 이름 (같은 이름이지만 대응을 명시)
FIELDS = [
    ("한전기준", "한전기준"),
    ("비고1", "비고1"),
    ("비고2", "비고2"),
    ("주소2", "주소2"),
]

# 필터 판정에만 쓰는 열(레코드에 붙이지는 않는다)
FILTER_COLS = ["주소", "대상"]

# 계기타입 재판정 — 계기번호 3~4번째 자리 코드**만** 본다(CLAUDE.md 데이터규칙).
#   값 표기는 이 데이터셋이 쓰는 어휘를 따른다(G타입 / AE타입 / E타입 / 보안계기).
#   ★A접두는 판정 근거가 아니다(영준님 정정 2026-08-18): "A접두는 아미고가 아님.
#     3,4번째 53 55 만 아미고임. 52도 아미고일 수 있으나 고압계기라 경험이 없을 뿐."
#     그래서 A0520004390(코드 52)은 코드 매핑 실패 -> 엑셀 값(G)을 그대로 따른다.
TYPE_BY_CODE = {
    "17": "E타입",
    "19": "AE타입",
    "25": "G타입", "26": "G타입", "27": "G타입",
    "45": "G타입", "46": "G타입", "47": "G타입",
    "53": "보안계기", "55": "보안계기",
}

# 매핑에 없는 코드(38·52 등)일 때 쓰는 엑셀 계기타입 표기 -> 데이터셋 어휘
TYPE_BY_XLS = {
    "E": "E타입",
    "AE": "AE타입", "EA": "AE타입",
    "G": "G타입",
    "AMIGO": "보안계기", "보안계기": "보안계기",
}

EMPTY_TOKENS = {"", "0", "#N/A", "#n/a", "nan", "NaN", "None", "none", "-"}


def norm(v):
    """셀 값 정규화 — 빈 의미의 토큰은 전부 빈 문자열로."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return "" if s in EMPTY_TOKENS else s


def meter_no(v):
    """계기번호 정규화 — float -> int -> str -> 하이픈 제거 -> zfill(11)."""
    if v is None:
        return ""
    if isinstance(v, float):
        v = int(v)
    s = str(v).strip().replace("-", "")
    if not s or s.lower() == "nan":
        return ""
    return s.zfill(11)


def load_sheet2(path=XLSX):
    """계기번호 -> {필드: 값} 사전."""
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header) if h}
    for xls_col in [c for c, _ in FIELDS] + FILTER_COLS + ["계기번호2", "계기타입"]:
        if xls_col not in idx:
            raise SystemExit("엑셀 {} 에 '{}' 열이 없다 — 시트를 확인하라".format(SHEET, xls_col))
    if "계기번호" not in idx:
        raise SystemExit("엑셀 {} 에 '계기번호' 열이 없다".format(SHEET))

    out = {}
    dups = []
    rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        no = meter_no(row[idx["계기번호"]])
        if not no or set(no) == {"0"}:
            continue
        rows += 1
        if no in out:
            dups.append(no)
        rec = {fld: norm(row[idx[col]]) for col, fld in FIELDS}
        # 필터 판정용 값은 '_' 를 붙여 따로 둔다 — 레코드에 붙는 필드(FIELDS)와 섞이지 않게.
        for col in FILTER_COLS:
            rec["_" + col] = norm(row[idx[col]])
        rec["_계기번호2"] = meter_no(row[idx["계기번호2"]])
        rec["_계기타입"] = norm(row[idx["계기타입"]])
        out[no] = rec
        # ★계기번호2 로도 찾을 수 있게 같은 레코드를 한 번 더 건다.
        #   교체를 이미 마친 파일에 이 스크립트를 다시 돌려도 매칭이 깨지지 않게(재실행 가능).
        n2 = rec["_계기번호2"]
        if n2 and n2 != no and n2 not in out:
            out[n2] = rec
    return out, rows, dups


def type_of(no):
    """계기번호 3~4번째 자리로 계기타입 판정. 매핑에 없으면 None(=코드로는 못 가림)."""
    return TYPE_BY_CODE.get(str(no or "")[2:4])


def apply_meter_no2(records, table):
    """계기번호를 계기번호2(현재 실제 계기번호)로 교체하고 계기타입을 재판정.

    영준님 2026-08-18: "계기교체 등으로 번호가 바뀐 게 있어 계기번호2 기준으로 리스트를
    꾸린다. 계기번호1은 데이터에도 디테일에도 남기지 않는다."
    교체 흔적은 레코드에 남기지 않고 별도 내역 파일로 뺀다.

    반환: (교체내역 리스트, 타입 판별불가 리스트)
    """
    changes = []
    unmapped = []
    for rec in records:
        old = str(rec.get("계기번호", "")).strip()
        src = table.get(old)
        if src is None:
            continue
        new = src.get("_계기번호2") or old
        old_type = rec.get("계기타입", "")
        new_type = type_of(new)
        if new_type is None:
            # 코드로 못 가리는 건(38·52 등) — 지어내지 않고 엑셀 Sheet2 계기타입을 그대로 쓴다.
            xls = src.get("_계기타입", "")
            new_type = TYPE_BY_XLS.get(xls.upper(), xls) or old_type
            unmapped.append((old, new, xls, new_type))
        if new != old or new_type != old_type:
            changes.append({
                "계기번호1": old,
                "계기번호2": new,
                "주소": rec.get("주소", ""),
                "지사": rec.get("지사", ""),
                "계기타입_변화": "{} -> {}".format(old_type or "(없음)", new_type or "(없음)"),
            })
        rec["계기번호"] = new
        rec["계기타입"] = new_type
    return changes, unmapped


def is_target(src):
    """작업대상 판정 — 한전기준 있음 + 주소 있음 + 대상=='철거' (셋 다 만족해야 대상)."""
    if src is None:
        return False
    return bool(src.get("한전기준")) and bool(src.get("_주소")) and src.get("_대상") == "철거"


def apply_to(records, table):
    """레코드에 필드 부착. 반환: (변경건수, 미매칭목록, 한전기준 분포)"""
    changed = 0
    unmatched = []
    dist = collections.Counter()
    for rec in records:
        no = str(rec.get("계기번호", "")).strip()
        src = table.get(no)
        if src is None:
            unmatched.append(no)
            dist["<미매칭>"] += 1
            continue
        touched = False
        for _, fld in FIELDS:
            val = src.get(fld, "")
            if rec.get(fld) != val:
                rec[fld] = val
                touched = True
        if touched:
            changed += 1
        dist[src.get("한전기준", "") or "<빈칸>"] += 1
    return changed, unmatched, dist


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="실제 파일에 반영(기본은 예행)")
    ap.add_argument("--no-filter", action="store_true",
                    help="작업대상 필터를 적용하지 않는다(필드만 부착)")
    ap.add_argument("--no-renumber", action="store_true",
                    help="계기번호2 교체·계기타입 재판정을 하지 않는다")
    ap.add_argument("--target", default=str(TARGET))
    args = ap.parse_args()

    target = Path(args.target)
    table, rows, dups = load_sheet2()
    print("[엑셀] {} {}행 읽음, 계기번호 유일 {}건, 중복 {}건".format(
        SHEET, rows, len(table), len(dups)))
    if dups:
        print("  중복 계기번호:", dups[:10])

    records = json.loads(target.read_text(encoding="utf-8"))
    changed, unmatched, dist = apply_to(records, table)

    print("[대상] {} {}건".format(target.name, len(records)))
    print("[매칭] 미매칭 {}건 {}".format(len(unmatched), unmatched[:10]))
    print("[한전기준 분포]")
    for k, v in sorted(dist.items(), key=lambda kv: -kv[1]):
        print("   {:<12} {}".format(k, v))
    print("[변경] {}건".format(changed))

    if unmatched:
        raise SystemExit("미매칭이 있다 — 반영을 중단한다")

    if not args.no_filter:
        kept = [r for r in records if is_target(table.get(str(r.get("계기번호", "")).strip()))]
        dropped = len(records) - len(kept)
        dist_keep = collections.Counter(r.get("한전기준", "") or "<빈칸>" for r in kept)
        print("[필터] 작업대상 {}건 유지 / {}건 제외".format(len(kept), dropped))
        for k, v in sorted(dist_keep.items(), key=lambda kv: -kv[1]):
            print("   {:<12} {}".format(k, v))
        print("[필터] 고유 주소 {}건".format(len({r.get("주소") for r in kept})))
        if dist_keep.get("<빈칸>"):
            raise SystemExit("필터 후에도 한전기준 빈칸이 남았다 — 판정을 확인하라")
        records = kept

    changes = []
    if not args.no_renumber:
        changes, unmapped = apply_meter_no2(records, table)
        nos = [r.get("계기번호", "") for r in records]
        print("[계기번호2] 교체·재판정 {}건 (번호 또는 타입이 바뀐 건)".format(len(changes)))
        print("   빈 번호 {}건 / 중복 {}건".format(
            sum(1 for n in nos if not n), len(nos) - len(set(nos))))
        for c in changes:
            print("   {} -> {} | {} | {}".format(
                c["계기번호1"], c["계기번호2"], c["계기타입_변화"], c["주소"]))
        if unmapped:
            print("[계기타입] 코드 매핑 없음 {}건 — 엑셀 Sheet2 값을 그대로 사용".format(len(unmapped)))
            for old, new, xls, t in unmapped:
                print("   {} -> {} (코드 {}) 엑셀 {} -> {}".format(
                    old, new, str(new)[2:4], xls or "(없음)", t or "(없음)"))
        if any(not n for n in nos) or len(nos) != len(set(nos)):
            raise SystemExit("계기번호 빈 값 또는 중복이 생겼다 — 반영을 중단한다")

    if not args.apply:
        print("\n예행 종료 — 반영하려면 --apply")
        return

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    if changes:
        # 계기번호1 은 레코드에 남기지 않는다(영준님 지시). 추적은 이 내역 파일로만 한다.
        hist = target.parent / "gapap_계기번호2교체내역_{}.json".format(stamp)
        hist.write_text(json.dumps(changes, ensure_ascii=False, indent=1), encoding="utf-8")
        print("[내역] {}".format(hist))
    backup = target.parent / "gapap-data.backup-적용전-{}.json".format(stamp)
    shutil.copy2(target, backup)
    print("[백업] {}".format(backup))

    # 원본이 한 줄 압축 JSON 이라 같은 형식으로 되쓴다(불필요한 전체 리포맷 방지).
    target.write_text(json.dumps(records, ensure_ascii=False), encoding="utf-8")
    print("[반영] {} 저장 완료".format(target))


if __name__ == "__main__":
    main()
