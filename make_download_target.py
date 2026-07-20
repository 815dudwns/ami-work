#!/usr/bin/env python3
"""
Download 폴더 사진 + 한전준공리스트 엑셀 매핑
출력: ocr_poc/download_target.csv
컬럼: 파일,계기번호,타입,검침값정답,fid
"""

import os
import re
import csv
import glob
import openpyxl

EXCEL_DIR = "/Users/woodelight/Projects/ami-work/data/종로_실효리스트_20260605"
DOWNLOAD_DIR = "/Users/woodelight/Projects/ami-work/data/phone_photos_20260606/Download/Download"
OUTPUT_CSV = "/Users/woodelight/Projects/ocr-meter/ocr_poc/download_target.csv"


def get_meter_type(meter_no: str) -> str:
    """계기번호 → 타입 판별. zfill(11)[2:4] 기준."""
    padded = meter_no.zfill(11)
    code = padded[2:4]
    type_map = {
        "17": "E",
        "19": "EA",
        "25": "G", "26": "G", "27": "G",
        "45": "G", "46": "G", "47": "G",
        "53": "Amigo", "55": "Amigo",
    }
    return type_map.get(code, "Unknown")


def build_excel_map(excel_dir: str) -> dict:
    """모든 xlsx에서 철거계기번호 → 철거지침 맵 구축."""
    meter_map = {}  # {계기번호: 검침값}
    xlsx_files = glob.glob(os.path.join(excel_dir, "*.xlsx"))
    print(f"엑셀 파일 {len(xlsx_files)}개 발견")

    for fpath in sorted(xlsx_files):
        fname = os.path.basename(fpath)
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active

        # 헤더 행 탐색 (철거전력량계번호 컬럼 위치 찾기)
        header_row = None
        col_remove_no = None
        col_remove_val = None
        col_install_no = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > 5:
                break
            for col_idx, cell in enumerate(row, start=1):
                if cell == "철거전력량계번호":
                    col_remove_no = col_idx
                    header_row = row_idx
                elif cell == "철거지침":
                    col_remove_val = col_idx
                elif cell == "부설전력량계번호":
                    col_install_no = col_idx

        if col_remove_no is None or col_remove_val is None:
            print(f"  [경고] {fname}: 철거 컬럼 미발견, 건너뜀")
            continue

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            remove_no = row[col_remove_no - 1]
            remove_val = row[col_remove_val - 1]
            if remove_no is None:
                continue
            remove_no_str = str(remove_no).strip().zfill(11)
            if remove_val is not None:
                meter_map[remove_no_str] = str(remove_val).strip()
                count += 1

        print(f"  {fname}: {count}건 로드")

    print(f"총 엑셀 맵 {len(meter_map)}건")
    return meter_map


def parse_filename(filename: str):
    """파일명에서 계기번호, 날짜, FID 추출.
    형식: {계기번호}_{YYYYMMDDHHMM}_{FID}.png.jpg
    """
    # 날짜 포함 패턴: 계기번호_날짜시간_FID.png.jpg
    m = re.match(r'^([^_]+)_(\d{12})_(.+)\.png\.jpg$', filename)
    if m:
        return m.group(1), m.group(2), m.group(3)
    return None, None, None


def main():
    # 1. 엑셀 맵 구축
    print("=== 엑셀 맵 구축 ===")
    meter_map = build_excel_map(EXCEL_DIR)

    # 2. Download 파일 목록
    print("\n=== Download 파일 처리 ===")
    jpg_files = sorted(f for f in os.listdir(DOWNLOAD_DIR) if f.endswith(".jpg"))
    print(f"전체 jpg 파일: {len(jpg_files)}개")

    rows = []
    matched = 0
    unmatched = 0
    parse_fail = 0

    for fname in jpg_files:
        meter_no, date_str, fid = parse_filename(fname)
        if meter_no is None:
            parse_fail += 1
            continue

        meter_no_padded = meter_no.zfill(11)
        meter_type = get_meter_type(meter_no_padded)
        full_path = os.path.join(DOWNLOAD_DIR, fname)

        if meter_no_padded in meter_map:
            answer = meter_map[meter_no_padded]
            matched += 1
        else:
            answer = ""
            unmatched += 1

        rows.append({
            "파일": full_path,
            "계기번호": meter_no_padded,
            "타입": meter_type,
            "검침값정답": answer,
            "fid": fid,
        })

    # 3. CSV 저장
    os.makedirs(os.path.dirname(OUTPUT_CSV), exist_ok=True)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["파일", "계기번호", "타입", "검침값정답", "fid"])
        writer.writeheader()
        writer.writerows(rows)

    # 4. 통계
    print(f"\n=== 결과 통계 ===")
    print(f"전체 jpg 파일:  {len(jpg_files)}")
    print(f"파싱 실패:      {parse_fail}")
    print(f"매핑 성공:      {matched}")
    print(f"미매핑:         {unmatched}")
    print(f"저장 경로:      {OUTPUT_CSV}")
    print(f"CSV 행 수:      {len(rows)}")

    # 타입별 통계
    type_counts = {}
    matched_rows = [r for r in rows if r["검침값정답"]]
    for r in matched_rows:
        t = r["타입"]
        type_counts[t] = type_counts.get(t, 0) + 1
    print(f"\n매핑 성공 타입별:")
    for t, c in sorted(type_counts.items()):
        print(f"  {t}: {c}")


if __name__ == "__main__":
    main()
