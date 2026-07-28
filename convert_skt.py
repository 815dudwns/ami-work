"""
SKT 중계기 AMI 모뎀 설치요청건 엑셀 → data/skt-data.json 변환
- 좌표 추출: 도로명 → 지번 → 동 폴백 (카카오 API)
- 비어있는 컬럼은 출력에 넣지 않음
- 계기위치세부 → 상호 필드로 매핑 (영준님 지시)
"""

import openpyxl
import json
import requests
import time
import re
import os
import sys

# 공용 자격증명 로더 (repo 루트 스크립트라 scripts/ 를 경로에 추가)
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scripts'))
from _env import require_env

API_KEY = require_env('KAKAO_REST_API_KEY')
EXCEL_FILE = '/Users/woodelight/Library/Mobile Documents/com~apple~CloudDocs/TalkFile_복사본 SKT 중계기 AMI 모뎀 설치요청건0411.xlsx'
OUTPUT_JSON = '/Users/woodelight/Projects/ami-work/data/skt-data.json'


def parse_type(meter_no):
    code = str(meter_no)[2:4]
    if code == '17': return 'E'
    if code == '19': return 'EA'
    if code in ('25', '26', '27', '45', '46', '47'): return 'G'
    if code in ('53', '55'): return 'Amigo'
    return '알수없음'


def fix_meter_no(val):
    if val is None: return ''.zfill(11)
    try:
        return str(int(float(val))).zfill(11)
    except (ValueError, TypeError):
        return str(val).strip().zfill(11)


def search_address(query):
    url = 'https://dapi.kakao.com/v2/local/search/address.json'
    headers = {'Authorization': f'KakaoAK {API_KEY}'}
    try:
        r = requests.get(url, headers=headers, params={'query': query}, timeout=10)
        if r.status_code == 200:
            docs = r.json().get('documents', [])
            if docs:
                doc = docs[0]
                lat = float(doc['y'])
                lng = float(doc['x'])
                road = (doc.get('road_address') or {}).get('address_name', '')
                jibun = (doc.get('address') or {}).get('address_name', '')
                road_name = road or jibun or query
                return road_name, lat, lng
    except Exception as e:
        print(f"  API 오류: {e}")
    return None


def extract_dong(address):
    m = re.match(r'(.*?[동읍면])', address)
    return m.group(1).strip() if m else None


def is_meaningful(v):
    """빈 문자열, None, '-' 같은 무의미한 값 제외"""
    if v is None: return False
    s = str(v).strip()
    if not s: return False
    if s in ('-', 'None', '#N/A', 'null', 'undefined'): return False
    return True


def safe_str(v):
    if not is_meaningful(v): return ''
    return str(v).strip()


def main():
    print("=" * 60)
    print("SKT 중계기 데이터 변환")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    print(f"컬럼: {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if not any(rec.values()): continue
        rows.append(rec)
    print(f"입력 행: {len(rows)}\n")

    result = []
    exact_cnt = approx_cnt = fail_cnt = 0

    for i, d in enumerate(rows, 1):
        jibun = safe_str(d.get('주소1'))
        road  = safe_str(d.get('주소2'))

        # 좌표 추출 — 도로명 → 지번 → 동 폴백
        found = None
        accuracy = None
        if road:
            found = search_address(road)
            if found: accuracy = 'exact'
        if not found and jibun:
            found = search_address(jibun)
            if found: accuracy = 'exact'
        if not found:
            dong = extract_dong(jibun) or extract_dong(road)
            if dong:
                found = search_address(dong)
                if found: accuracy = 'approximate'

        if not found:
            print(f"[{i}/{len(rows)}] FAIL: {jibun or road}")
            fail_cnt += 1
            time.sleep(0.15)
            continue

        resolved_road, lat, lng = found
        if accuracy == 'exact': exact_cnt += 1
        else: approx_cnt += 1

        meter_no = fix_meter_no(d.get('계기번호'))
        # 계기위치세부 → 상호 (영준님 지시)
        sangho = safe_str(d.get('계기위치세부'))

        rec = {
            '계기번호': meter_no,
            '계기타입': parse_type(meter_no),
            '주소': jibun,
            '도로명주소': road or resolved_road,
            'lat': lat,
            'lng': lng,
            '좌표정확도': accuracy,
        }
        # 채워진 값만 추가
        if safe_str(d.get('지사')):       rec['지사']     = safe_str(d.get('지사'))
        if safe_str(d.get('고객번호')):    rec['고객번호'] = safe_str(d.get('고객번호'))
        if safe_str(d.get('변대번호')):    rec['변대주']   = safe_str(d.get('변대번호'))
        if sangho:                          rec['상호']     = sangho
        if is_meaningful(d.get('실효\n년월')):
            rec['실효년월'] = safe_str(d.get('실효\n년월'))

        # SKT 전용 메타 (값 있을 때만)
        if is_meaningful(d.get('KDN작업일')):
            rec['skt_kdn_작업일'] = safe_str(d.get('KDN작업일'))
        if is_meaningful(d.get('작업결과\n(O/X)')):
            rec['skt_작업결과'] = safe_str(d.get('작업결과\n(O/X)'))
        if is_meaningful(d.get('불가사유')):
            rec['skt_불가사유'] = safe_str(d.get('불가사유'))
        if is_meaningful(d.get('비고')):
            rec['skt_비고'] = safe_str(d.get('비고'))
        if is_meaningful(d.get('KDN\n작업이력')):
            rec['skt_kdn_이력'] = safe_str(d.get('KDN\n작업이력'))

        result.append(rec)
        flag = 'OK' if accuracy == 'exact' else '~ '
        print(f"[{i}/{len(rows)}] {flag} {jibun}")
        time.sleep(0.15)

    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print()
    print("=" * 60)
    print(f"exact:       {exact_cnt}")
    print(f"approximate: {approx_cnt}")
    print(f"fail:        {fail_cnt}")
    print(f"total saved: {len(result)}")
    print(f"=> {OUTPUT_JSON}")
    print("=" * 60)


if __name__ == '__main__':
    main()
